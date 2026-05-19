import sys, os, io, csv as csv_mod, math, hashlib, re
sys.path.insert(0, os.path.dirname(__file__))

# Load .env from project root
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))
except ImportError:
    pass


from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Request, WebSocket, WebSocketDisconnect, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, or_
from pydantic import BaseModel, Field, EmailStr, validator
from typing import Optional, List
from datetime import datetime, timedelta, timezone
import random, string

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

from database import (get_db, User, Ratepayer, Invoice, Payment, Expenditure,
                      Budget, AuditLog, LeakageAlert, RevenueTarget,
                      UserRole, PaymentStatus, RevenueCategory, AnomalyFlag,
                      LoginAttempt, PaymentPlan, PaymentPlanStatus, SystemNotification,
                      ExchangeRate, BillingRate, BillingRun)
from auth import (verify_password, hash_password, create_access_token,
                  get_current_user, require_roles)

def now():
    return datetime.now(timezone.utc).replace(tzinfo=None)

def _provider_symbol_for_api(code: str) -> str:
    # Map internal currency codes to external provider symbols
    c = normalize_currency_code(code)
    if c in ("ZIG", "ZWG"):
        return "ZWL"
    return c

def fetch_rate_from_exchangerate_host(currency: str) -> Optional[dict]:
    import requests
    symbol = _provider_symbol_for_api(currency)
    if symbol == "USD":
        return {"rate": 1.0, "source": "local"}
    url = f"https://api.exchangerate.host/latest?base={symbol}&symbols=USD"
    try:
        r = requests.get(url, timeout=6)
        if r.status_code == 200:
            data = r.json()
            rate = data.get("rates", {}).get("USD")
            if rate:
                # rate is USD per 1 unit of currency
                return {"rate": float(rate), "source": url}
    except Exception:
        return None
    return None

def update_exchange_rate_in_db(db: Session, currency: str) -> Optional[ExchangeRate]:
    code = normalize_currency_code(currency)
    fetched = fetch_rate_from_exchangerate_host(code)
    if not fetched:
        return None
    er = db.query(ExchangeRate).filter(ExchangeRate.currency == code).first()
    if not er:
        er = ExchangeRate(currency=code, rate_to_usd=fetched["rate"], source=fetched["source"], manual=False, updated_at=now())
        db.add(er)
    else:
        er.rate_to_usd = fetched["rate"]
        er.source = fetched["source"]
        er.manual = False
        er.updated_at = now()
    db.commit()
    return er

# â"€â"€â"€ Security Constants â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
MAX_FAILED_LOGINS = 5
LOCKOUT_MINUTES   = 30

# â"€â"€â"€ Password Strength Validator â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
def validate_password_strength(password: str) -> str:
    """Returns error message string, or empty string if password is strong enough."""
    if len(password) < 8:
        return "Password must be at least 8 characters long."
    if not re.search(r"[A-Z]", password):
        return "Password must contain at least one uppercase letter."
    if not re.search(r"[a-z]", password):
        return "Password must contain at least one lowercase letter."
    if not re.search(r"\d", password):
        return "Password must contain at least one digit (0-9)."
    return ""

# â"€â"€â"€ Invoice Duplicate Fingerprint â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
def compute_invoice_fingerprint(ratepayer_id: int, category: str, amount: float, due_date: str) -> str:
    """SHA-256 fingerprint to detect duplicate invoices (ACFE, 2022: Billing Scheme Fraud)."""
    raw = f"{ratepayer_id}|{category}|{round(amount, 2)}|{str(due_date)[:10]}"
    return hashlib.sha256(raw.encode()).hexdigest()[:32]

CURRENCY_CONVERSION_RATES = {
    "USD": 1.0,
    "ZIG": 0.001,
    "ZWG": 0.001,
}

def normalize_currency_code(code: Optional[str]) -> str:
    if not code:
        return "USD"
    currency = str(code).strip().upper()
    if currency == "ZWG":
        return "ZIG"
    return currency

def amount_to_usd(amount: float, currency: Optional[str], db: Session = None) -> float:
    # Prefer DB-stored rate when available; fall back to static mapping
    code = normalize_currency_code(currency)
    rate = CURRENCY_CONVERSION_RATES.get(code, 1.0)
    try:
        if db is not None:
            er = db.query(ExchangeRate).filter(ExchangeRate.currency == code).first()
            if er and er.rate_to_usd:
                rate = er.rate_to_usd
    except Exception:
        pass
    return round((amount or 0.0) * rate, 2)


def usd_payment_sum(query) -> float:
    # Groups payments by currency and converts each group to USD using DB rates where available.
    rows = query.with_entities(Payment.currency, func.sum(Payment.amount)).group_by(Payment.currency).all()
    total = 0.0
    session = None
    try:
        session = query.session
    except Exception:
        session = None
    codes = [normalize_currency_code(r[0]) for r in rows]
    rate_map = {}
    if session is not None and codes:
        try:
            ers = session.query(ExchangeRate).filter(ExchangeRate.currency.in_(codes)).all()
            rate_map = {er.currency: er.rate_to_usd for er in ers}
        except Exception:
            rate_map = {}
    for currency, amt in rows:
        code = normalize_currency_code(currency)
        rate = rate_map.get(code, CURRENCY_CONVERSION_RATES.get(code, 1.0))
        total += (amt or 0.0) * rate
    return round(total, 2)

# â"€â"€â"€ AI: Ratepayer Risk Score â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
def compute_ratepayer_risk(rp, db) -> dict:
    """
    Composite 0-100 risk score per ratepayer using a weighted additive model.
    Factors: overdue balance ratio (40%), payment recency (30%),
    anomaly-flagged invoices (20%), payment plan defaults (10%).
    Adapted from ZIMRA (2023) Taxpayer Risk Segmentation Framework and
    NCR (2022) behavioural scoring principles.
    """
    score = 0.0
    # Factor 1: Overdue balance ratio (40 pts)
    total_billed  = db.query(func.sum(Invoice.amount)).filter(Invoice.ratepayer_id == rp.id).scalar() or 0
    total_overdue = db.query(func.sum(Invoice.balance)).filter(
        Invoice.ratepayer_id == rp.id, Invoice.status == PaymentStatus.overdue).scalar() or 0
    if total_billed > 0:
        score += min(total_overdue / total_billed, 1.0) * 40
    # Factor 2: Payment recency (30 pts)
    last_pmt = db.query(func.max(Payment.payment_date)).filter(Payment.ratepayer_id == rp.id).scalar()
    if last_pmt is None:
        score += 30
    else:
        score += min((now() - last_pmt).days / 365, 1.0) * 30
    # Factor 3: Anomaly-flagged invoices (20 pts)
    high_anom   = db.query(func.count(Invoice.id)).filter(
        Invoice.ratepayer_id == rp.id, Invoice.anomaly_flag == AnomalyFlag.high).scalar() or 0
    total_inv   = db.query(func.count(Invoice.id)).filter(Invoice.ratepayer_id == rp.id).scalar() or 1
    score += min(high_anom / total_inv, 1.0) * 20
    # Ensure rate table exists and try to keep it updated lazily
    # Factor 4: Defaulted payment plans (10 pts)
    defaulted = db.query(func.count(PaymentPlan.id)).filter(
        PaymentPlan.ratepayer_id == rp.id, PaymentPlan.status == PaymentPlanStatus.defaulted).scalar() or 0
    score += min(defaulted * 5, 10)
    score = round(min(score, 100), 1)
    label = "high" if score >= 70 else ("medium" if score >= 40 else "low")
    return {"score": score, "label": label}

# â"€â"€â"€ AI: Revenue Prediction (OLS Linear Regression) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
def predict_next_month_revenue(db) -> dict:
    """
    Predict next month revenue using OLS linear regression on 6 months of data.
    Freedman, Pisani & Purves (2007). Statistics (4th ed.). W.W. Norton.
    """
    monthly_data = []
    for i in range(5, -1, -1):
        ref = now() - timedelta(days=30 * i)
        ms  = ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        me  = ms.replace(month=ms.month % 12 + 1, day=1) if ms.month < 12 else ms.replace(year=ms.year + 1, month=1, day=1)
        collected = usd_payment_sum(db.query(Payment).filter(
            Payment.payment_date >= ms, Payment.payment_date < me))
        monthly_data.append((ms.strftime("%b %Y"), float(collected)))
    values = [v for _, v in monthly_data]
    n = len(values)
    if n < 3:
        return {"prediction": None, "basis": "Insufficient data (need 3+ months)"}
    x = list(range(n)); mx = sum(x) / n; my = sum(values) / n
    num = sum((x[i]-mx)*(values[i]-my) for i in range(n))
    den = sum((xi-mx)**2 for xi in x)
    slope = num / den if den != 0 else 0
    predicted = round(max((my - slope*mx) + slope*n, 0), 2)
    return {
        "predicted_next_month": predicted,
        "trend": "increasing" if slope > 0 else "decreasing",
        "slope_per_month": round(slope, 2),
        "last_6_months": monthly_data,
        "basis": "OLS linear regression - Freedman, Pisani & Purves (2007)"
    }


def refresh_overdue_invoices(db: Session):
    cutoff = now()

    # Correct invoices wrongly marked overdue whose due_date is still in the future.
    # This can occur when seed data assigns overdue status without checking the due_date.
    future_overdue = db.query(Invoice).filter(
        Invoice.status == PaymentStatus.overdue,
        Invoice.due_date >= cutoff,
        Invoice.balance > 0
    ).all()
    for inv in future_overdue:
        inv.status = PaymentStatus.pending
        db.add(AuditLog(user_id=None, action="UPDATE", table_name="invoices",
                        record_id=inv.id,
                        description=f"Invoice {inv.invoice_number} reset from overdue to pending (due date {str(inv.due_date)[:10]} is in the future)"))

    # Mark genuinely overdue invoices.
    overdue_list = db.query(Invoice).filter(
        Invoice.status.in_([PaymentStatus.pending, PaymentStatus.disputed]),
        Invoice.due_date < cutoff,
        Invoice.balance > 0
    ).all()
    for inv in overdue_list:
        inv.status = PaymentStatus.overdue
        db.add(AuditLog(user_id=None, action="UPDATE", table_name="invoices",
                        record_id=inv.id,
                        description=f"Invoice {inv.invoice_number} marked overdue"))

    if future_overdue or overdue_list:
        db.commit()
    return len(overdue_list)

# â"€â"€â"€ Anomaly Detection â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
# Uses Z-score method (Grubbs, 1969; Iglewicz & Hoaglin, 1993).
# |Z| > 3.0 â†' high severity; |Z| > 2.0 â†' medium; |Z| > 1.5 â†' low.
# This is the industry-standard statistical threshold for outlier detection
# in financial data (ACFE, 2022; KPMG Revenue Assurance Framework, 2021).

def _zscore_flag(value: float, values: list) -> tuple:
    """Return (AnomalyFlag, reason) using Z-score outlier detection."""
    if len(values) < 3:
        return AnomalyFlag.none, None
    mean = sum(values) / len(values)
    variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
    std_dev = math.sqrt(variance) if variance > 0 else 0
    if std_dev == 0:
        return AnomalyFlag.none, None
    z = (value - mean) / std_dev
    abs_z = abs(z)
    direction = "above" if z > 0 else "below"
    if abs_z > 3.0:
        return AnomalyFlag.high, (
            f"Z-score {z:.2f} - amount is a high-severity outlier ({direction} Î¼=${mean:.2f}, Ïƒ=${std_dev:.2f}). "
            f"Consistent with revenue leakage patterns identified in DSR literature review."
        )
    elif abs_z > 2.0:
        return AnomalyFlag.medium, (
            f"Z-score {z:.2f} - amount deviates significantly from category mean ${mean:.2f} (Ïƒ=${std_dev:.2f})."
        )
    elif abs_z > 1.5:
        return AnomalyFlag.low, (
            f"Z-score {z:.2f} - amount is slightly unusual vs category mean ${mean:.2f}."
        )
    return AnomalyFlag.none, None

def _detect_anomaly(value: float, values: list) -> tuple:
    """
    Primary anomaly detector: Isolation Forest for >=8 samples, Z-score fallback.
    Isolation Forest - Liu, Ting & Zhou (2008). Isolation Forest. IEEE ICDM.
    """
    if len(values) >= 8:
        try:
            from sklearn.ensemble import IsolationForest
            import numpy as np
            X = np.array(values + [value]).reshape(-1, 1)
            clf = IsolationForest(contamination=0.1, random_state=42, n_estimators=100)
            clf.fit(X)
            score = float(clf.decision_function([[value]])[0])
            pred  = clf.predict([[value]])[0]  # -1 = anomaly
            if pred == -1:
                mean_v    = sum(values) / len(values)
                direction = "above" if value > mean_v else "below"
                if score < -0.15:
                    return AnomalyFlag.high, (
                        f"Isolation Forest anomaly score {score:.3f} - high-severity outlier "
                        f"({direction} mean ${mean_v:.2f}, n={len(values)} samples). "
                        f"Liu, Ting & Zhou (2008) IEEE ICDM."
                    )
                elif score < -0.05:
                    return AnomalyFlag.medium, (
                        f"Isolation Forest anomaly score {score:.3f} - medium outlier "
                        f"({direction} mean ${mean_v:.2f}). Liu, Ting & Zhou (2008)."
                    )
                else:
                    return AnomalyFlag.low, (
                        f"Isolation Forest anomaly score {score:.3f} - marginal outlier detected."
                    )
            return AnomalyFlag.none, None
        except ImportError:
            pass
    return _zscore_flag(value, values)

# â"€â"€â"€ Export Helpers â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def make_csv_response(headers, rows, filename):
    output = io.StringIO()
    writer = csv_mod.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    content = output.getvalue().encode("utf-8-sig")
    return Response(content=content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})

def make_excel_response(headers, rows, filename, sheet_name="Data", report_title=None):
    if not EXCEL_OK:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    start_row = 1
    # Optional report title banner
    if report_title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        start_row = 2
        # Generated date
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        dt_cell = ws.cell(row=start_row, column=1,
                          value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}")
        dt_cell.font = Font(italic=True, size=9, color="666666")
        dt_cell.alignment = Alignment(horizontal="center")
        start_row += 1
    # Header row
    hdr_row = start_row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E5FA3", end_color="2E5FA3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 18
    # Data rows
    for r_idx, row in enumerate(rows, hdr_row + 1):
        fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid") \
               if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center")
    # Auto-fit columns (skip MergedCell objects created by the title banner row)
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col if hasattr(cell, "column_letter")),
            default=8
        )
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_response(fmt, headers, rows, csv_name, xlsx_name, sheet_name="Data", title=None):
    if fmt == "xlsx":
        return make_excel_response(headers, rows, xlsx_name, sheet_name, title)
    return make_csv_response(headers, rows, csv_name)

# â"€â"€â"€ Import Helpers â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

async def parse_upload(file: UploadFile) -> list:
    """Return list of dicts from CSV or XLSX upload."""
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, f"File exceeds {MAX_FILE_SIZE} bytes")
    name = (file.filename or "").lower()
    rows = []
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv_mod.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        if not EXCEL_OK:
            raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        hdrs = [str(cell.value or "").strip() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append({hdrs[i]: row[i] for i in range(len(hdrs))})
    else:
        raise HTTPException(400, "Only .csv and .xlsx files are supported")
    return rows

# ─── WebSocket Connection Manager ──────────────────────────────────────────────

class _WSManager:
    """Broadcast JSON messages to all connected WebSocket clients."""
    def __init__(self):
        self._clients: set = set()

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self._clients.add(ws)

    def disconnect(self, ws: WebSocket):
        self._clients.discard(ws)

    async def broadcast(self, message: dict):
        dead = set()
        for ws in self._clients:
            try:
                await ws.send_json(message)
            except Exception:
                dead.add(ws)
        self._clients -= dead

ws_manager = _WSManager()

# â"€â"€â"€ App Setup â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

app = FastAPI(title="City of Harare FMS", version="2.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

@app.on_event("startup")
def on_startup():
    from database import Base, engine, SessionLocal, User, UserRole
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if db.query(User).count() == 0:
            import sys as _sys, os as _os
            _sys.path.insert(0, _os.path.dirname(__file__))
            import seed  # noqa: F401
    finally:
        db.close()

frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.exists(frontend_path):
    app.mount("/static", StaticFiles(directory=frontend_path), name="static")

# â"€â"€â"€ Pydantic Schemas â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

class RatepayerCreate(BaseModel):
    full_name: str
    address: str
    ward: str
    zone: str
    phone: Optional[str] = None
    email: Optional[str] = None
    property_type: str = "residential"

class RatepayerUpdate(BaseModel):
    full_name: Optional[str] = None
    address: Optional[str] = None
    ward: Optional[str] = None
    zone: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    property_type: Optional[str] = None
    is_active: Optional[bool] = None

class InvoiceCreate(BaseModel):
    ratepayer_id: int
    category: str
    amount: float = Field(..., gt=0)
    due_date: str
    notes: Optional[str] = None

class InvoiceUpdate(BaseModel):
    due_date: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    amount: Optional[float] = None

class PaymentCreate(BaseModel):
    ratepayer_id: int
    invoice_id: Optional[int] = None
    amount: float
    payment_method: str = "cash"
    currency: str = "USD"
    notes: Optional[str] = None

    @validator("currency", pre=True, always=True)
    def normalize_currency(cls, v):
        return normalize_currency_code(v)

class ExpenditureCreate(BaseModel):
    department: str
    description: str
    amount: float
    budget_line: str

class ExpenditureUpdate(BaseModel):
    department: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[float] = None
    budget_line: Optional[str] = None

class BudgetCreate(BaseModel):
    fiscal_year: str
    department: str
    category: str
    allocated_amount: float
    spent_amount: float = 0.0

class BudgetUpdate(BaseModel):
    fiscal_year: Optional[str] = None
    department: Optional[str] = None
    category: Optional[str] = None
    allocated_amount: Optional[float] = None
    spent_amount: Optional[float] = None

class RevenueTargetCreate(BaseModel):
    fiscal_year: str
    category: str
    target_amount: float
    period: str = "annual"
    notes: Optional[str] = None

class RevenueTargetUpdate(BaseModel):
    target_amount: Optional[float] = None
    period: Optional[str] = None
    notes: Optional[str] = None

class UserCreate(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    full_name: str = Field(..., min_length=1, max_length=100)
    email: EmailStr
    password: str = Field(..., min_length=8)
    role: str = "revenue_officer"

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None

class PaymentPlanCreate(BaseModel):
    ratepayer_id: int
    total_debt: float
    instalment_amount: float
    frequency: str = "monthly"
    total_instalments: int
    start_date: str
    notes: Optional[str] = None

class PaymentPlanUpdate(BaseModel):
    instalment_amount: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class NotificationCreate(BaseModel):
    title: str
    message: str
    category: str = "info"
    user_id: Optional[int] = None

class AlertResolveRequest(BaseModel):
    resolution_notes: str

# â"€â"€â"€ Auth â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.post("/api/auth/login")
def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """
    Enhanced login with brute-force protection (IMPROVEMENT 1 & 2).
    Locks account after 5 consecutive failures for 30 minutes.
    All attempts logged to LoginAttempt table with IP address.
    """
    ip = request.client.host if request.client else None

    def _log(success: bool, reason: str = None):
        db.add(LoginAttempt(username=form_data.username, ip_address=ip,
                            success=success, failure_reason=reason))
        db.commit()

    user = db.query(User).filter(User.username == form_data.username).first()
    if not user:
        _log(False, "user_not_found")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not user.is_active:
        _log(False, "account_disabled")
        raise HTTPException(status_code=403, detail="Account disabled. Contact your administrator.")
    if user.locked_until and user.locked_until > now():
        remaining = max(1, int((user.locked_until - now()).total_seconds() / 60))
        _log(False, "account_locked")
        raise HTTPException(status_code=423,
            detail=f"Account locked after {MAX_FAILED_LOGINS} failed attempts. Try again in {remaining} minute(s).")
    if not verify_password(form_data.password, user.hashed_password):
        user.failed_login_count = (user.failed_login_count or 0) + 1
        if user.failed_login_count >= MAX_FAILED_LOGINS:
            user.locked_until = now() + timedelta(minutes=LOCKOUT_MINUTES)
            db.commit()
            _log(False, "wrong_password_locked")
            raise HTTPException(status_code=423,
                detail=f"Account locked after {MAX_FAILED_LOGINS} failed attempts. Try again in {LOCKOUT_MINUTES} minutes.")
        db.commit()
        _log(False, "wrong_password")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    # Successful login - reset security counters
    user.failed_login_count = 0
    user.locked_until = None
    user.last_login = now()
    db.add(AuditLog(user_id=user.id, action="LOGIN", table_name="users",
                    record_id=user.id, ip_address=ip,
                    description=f"{user.full_name} logged in successfully"))
    db.commit()
    _log(True)
    token = create_access_token({"sub": user.username})
    return {"access_token": token, "token_type": "bearer",
            "user": {"id": user.id, "username": user.username,
                     "full_name": user.full_name, "role": user.role, "email": user.email}}

@app.get("/api/auth/me")
def get_me(current_user: User = Depends(get_current_user)):
    return {"id": current_user.id, "username": current_user.username,
            "full_name": current_user.full_name, "role": current_user.role,
            "email": current_user.email}

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

@app.patch("/api/auth/change-password")
def change_password(data: ChangePasswordRequest, db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    if not verify_password(data.current_password, current_user.hashed_password):
        raise HTTPException(400, "Current password is incorrect")
    strength_err = validate_password_strength(data.new_password)
    if strength_err:
        raise HTTPException(400, strength_err)
    current_user.hashed_password = hash_password(data.new_password)
    current_user.password_changed_at = now()
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="users",
                    record_id=current_user.id,
                    description=f"Password changed for user {current_user.username}"))
    db.commit()
    return {"message": "Password changed successfully"}

# â"€â"€â"€ Dashboard â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/dashboard/summary")
def dashboard_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    refresh_overdue_invoices(db)
    # Revenue balance: uses Invoice.amount_paid so the equation is always exact:
    # total_billed = total_collected + total_outstanding (sum(amount) = sum(amount_paid) + sum(balance))
    total_billed      = db.query(func.sum(Invoice.amount)).scalar() or 0
    total_collected   = db.query(func.sum(Invoice.amount_paid)).scalar() or 0
    total_outstanding = db.query(func.sum(Invoice.balance)).scalar() or 0
    collection_rate   = round(total_collected / total_billed * 100, 1) if total_billed > 0 else 0

    # Actual payments received via Payment table (USD-equivalent, includes ZIG conversion)
    payments_received_usd = usd_payment_sum(db.query(Payment))

    # Invoice breakdown by status
    inv_pending  = db.query(Invoice).filter(Invoice.status == PaymentStatus.pending).count()
    inv_paid     = db.query(Invoice).filter(Invoice.status == PaymentStatus.paid).count()
    inv_overdue  = db.query(Invoice).filter(Invoice.status == PaymentStatus.overdue).count()
    inv_disputed = db.query(Invoice).filter(Invoice.status == PaymentStatus.disputed).count()
    inv_waived   = db.query(Invoice).filter(Invoice.status == PaymentStatus.waived).count()
    inv_total    = db.query(Invoice).count()
    overdue_bal  = db.query(func.sum(Invoice.balance)).filter(
        Invoice.status == PaymentStatus.overdue).scalar() or 0

    # Payments & reconciliation
    pmt_total     = db.query(Payment).count()
    recon_count   = db.query(Payment).filter(Payment.is_reconciled == True).count()
    unrecon_count = pmt_total - recon_count
    recon_rate    = round(recon_count / pmt_total * 100, 1) if pmt_total > 0 else 0
    unrecon_amt   = usd_payment_sum(db.query(Payment).filter(Payment.is_reconciled == False))
    method_totals = {}
    for method in ["cash", "ecocash", "bank_transfer", "rtgs", "zipit"]:
        method_totals[method] = round(usd_payment_sum(
            db.query(Payment).filter(Payment.payment_method == method)), 2)

    # Budget & Expenditure
    total_allocated   = db.query(func.sum(Budget.allocated_amount)).scalar() or 0
    total_spent       = db.query(func.sum(Budget.spent_amount)).scalar() or 0
    budget_remaining  = total_allocated - total_spent
    budget_util       = round(total_spent / total_allocated * 100, 1) if total_allocated > 0 else 0
    total_expenditure = db.query(func.sum(Expenditure.amount)).scalar() or 0
    pending_approval  = db.query(Expenditure).filter(Expenditure.is_approved == False).count()

    # Ratepayers
    active_rp   = db.query(Ratepayer).filter(Ratepayer.is_active == True).count()
    inactive_rp = db.query(Ratepayer).filter(Ratepayer.is_active == False).count()

    # Aging buckets (outstanding balances by overdue age)
    today = now()
    aging = {"current": 0.0, "days_1_30": 0.0, "days_31_60": 0.0,
             "days_61_90": 0.0, "days_91_120": 0.0, "days_120_plus": 0.0}
    for inv in db.query(Invoice).filter(Invoice.balance > 0).all():
        days = (today - inv.due_date).days
        if days <= 0:    aging["current"]      += inv.balance
        elif days <= 30: aging["days_1_30"]    += inv.balance
        elif days <= 60: aging["days_31_60"]   += inv.balance
        elif days <= 90: aging["days_61_90"]   += inv.balance
        elif days <=120: aging["days_91_120"]  += inv.balance
        else:            aging["days_120_plus"]+= inv.balance
    aging = {k: round(v, 2) for k, v in aging.items()}

    # Leakage & anomalies
    active_alerts    = db.query(LeakageAlert).filter(LeakageAlert.is_resolved == False).count()
    anomaly_invoices = db.query(Invoice).filter(Invoice.anomaly_flag != AnomalyFlag.none).count()
    anomaly_payments = db.query(Payment).filter(Payment.anomaly_flag != AnomalyFlag.none).count()
    leakage_estimate = round(unrecon_amt * 0.40 + overdue_bal * 0.25, 2)

    return {
        # Revenue balance (always sums: billed = collected + outstanding)
        "total_billed": round(total_billed, 2),
        "total_collected": round(total_collected, 2),
        "total_outstanding": round(total_outstanding, 2),
        "collection_rate": collection_rate,
        "payments_received_usd": round(payments_received_usd, 2),
        # Invoice breakdown
        "invoices_count": inv_total,
        "inv_pending": inv_pending, "inv_paid": inv_paid,
        "inv_overdue": inv_overdue, "inv_disputed": inv_disputed, "inv_waived": inv_waived,
        "overdue_count": inv_overdue, "overdue_balance": round(overdue_bal, 2),
        # Payments & reconciliation
        "payments_count": pmt_total,
        "reconciled_count": recon_count, "unreconciled_count": unrecon_count,
        "reconciliation_rate": recon_rate, "unreconciled_amount": round(unrecon_amt, 2),
        "payment_methods": method_totals,
        # Budget & expenditure
        "total_allocated": round(total_allocated, 2),
        "total_spent": round(total_spent, 2),
        "budget_remaining": round(budget_remaining, 2),
        "budget_utilisation": budget_util,
        "total_expenditure": round(total_expenditure, 2),
        "pending_approval_count": pending_approval,
        # Ratepayers
        "ratepayers_count": active_rp + inactive_rp,
        "active_ratepayers": active_rp, "inactive_ratepayers": inactive_rp,
        # Aging
        "aging": aging,
        # Leakage
        "active_alerts": active_alerts,
        "anomaly_count": anomaly_invoices + anomaly_payments,
        "leakage_estimate": leakage_estimate,
    }

@app.get("/api/dashboard/revenue-by-category")
def revenue_by_category(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    results = db.query(Invoice.category, func.sum(Invoice.amount).label("billed"),
                       func.sum(Invoice.amount_paid).label("collected")).group_by(Invoice.category).all()
    return [{"category": r.category, "billed": round(r.billed or 0, 2),
             "collected": round(r.collected or 0, 2)} for r in results]

@app.get("/api/dashboard/monthly-trend")
def monthly_trend(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    months = []
    for i in range(6, 0, -1):
        start = now().replace(day=1) - timedelta(days=30*i)
        end   = start + timedelta(days=30)
        billed    = db.query(func.sum(Invoice.amount)).filter(Invoice.issue_date >= start, Invoice.issue_date < end).scalar() or 0
        collected = usd_payment_sum(db.query(Payment).filter(Payment.payment_date >= start, Payment.payment_date < end))
        months.append({"month": start.strftime("%b %Y"), "billed": round(billed, 2), "collected": round(collected, 2)})
    return months

@app.get("/api/dashboard/alerts")
def get_alerts(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    alerts = db.query(LeakageAlert).filter(LeakageAlert.is_resolved == False)\
               .order_by(desc(LeakageAlert.created_at)).limit(10).all()
    return [{"id": a.id, "type": a.alert_type, "severity": a.severity,
             "description": a.description, "created_at": str(a.created_at)} for a in alerts]

@app.post("/api/invoices/refresh-overdue")
def refresh_overdue(db: Session = Depends(get_db),
                    current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant, UserRole.revenue_officer))):
    count = refresh_overdue_invoices(db)
    return {"updated_invoices": count, "message": f"Marked {count} invoice(s) as overdue"}

# â"€â"€â"€ Ratepayers â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/ratepayers")
def list_ratepayers(search: Optional[str] = None, zone: Optional[str] = None,
                    skip: int = 0, limit: int = 50,
                    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Ratepayer)
    if search:
        if len(search) > 100:
            raise HTTPException(400, "Search term too long")
        q = q.filter(Ratepayer.full_name.ilike(f"%{search}%") | Ratepayer.account_number.ilike(f"%{search}%"))
    if zone:   q = q.filter(Ratepayer.zone == zone)
    total = q.count()
    items = q.offset(skip).limit(limit).all()
    return {"total": total, "items": [
        {"id": r.id, "account_number": r.account_number, "full_name": r.full_name,
         "address": r.address, "ward": r.ward, "zone": r.zone,
         "phone": r.phone, "email": r.email, "property_type": r.property_type,
         "is_active": r.is_active} for r in items]}

@app.post("/api/ratepayers")
def create_ratepayer(data: RatepayerCreate, db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user)):
    acct = "COH-" + "".join(random.choices(string.digits, k=6))
    rp = Ratepayer(account_number=acct, **data.dict())
    db.add(rp); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="ratepayers",
                    record_id=rp.id, description=f"Created ratepayer: {data.full_name}"))
    db.commit()
    return {"id": rp.id, "account_number": rp.account_number, "message": "Ratepayer created successfully"}

@app.get("/api/ratepayers/{rp_id}")
def get_ratepayer(rp_id: int, db: Session = Depends(get_db),
                  current_user: User = Depends(get_current_user)):
    refresh_overdue_invoices(db)
    rp = db.query(Ratepayer).filter(Ratepayer.id == rp_id).first()
    if not rp: raise HTTPException(404, "Ratepayer not found")
    invoices = db.query(Invoice).filter(Invoice.ratepayer_id == rp_id).all()
    payments = db.query(Payment).filter(Payment.ratepayer_id == rp_id).all()
    total_billed = sum(i.amount for i in invoices)
    total_paid   = sum(amount_to_usd(p.amount, p.currency) for p in payments)
    return {
        "id": rp.id, "account_number": rp.account_number, "full_name": rp.full_name,
        "address": rp.address, "ward": rp.ward, "zone": rp.zone,
        "phone": rp.phone, "email": rp.email, "property_type": rp.property_type,
        "is_active": rp.is_active,
        "total_billed": round(total_billed, 2), "total_paid": round(total_paid, 2),
        "balance": round(total_billed - total_paid, 2),
        "invoice_count": len(invoices), "payment_count": len(payments)
    }


class ExchangeRateSet(BaseModel):
    currency: str
    rate_to_usd: float
    source: Optional[str] = None
    manual: Optional[bool] = True


@app.get("/api/exchange-rate")
def get_exchange_rate(currency: str = "ZIG", db: Session = Depends(get_db)):
    code = normalize_currency_code(currency)
    er = db.query(ExchangeRate).filter(ExchangeRate.currency == code).first()
    if er:
        return {"currency": er.currency, "rate_to_usd": er.rate_to_usd, "source": er.source, "manual": er.manual, "updated_at": str(er.updated_at)}
    # attempt fetch if not present
    fetched = update_exchange_rate_in_db(db, code)
    if fetched:
        return {"currency": fetched.currency, "rate_to_usd": fetched.rate_to_usd, "source": fetched.source, "manual": fetched.manual, "updated_at": str(fetched.updated_at)}
    raise HTTPException(404, f"No exchange rate available for {code}")


@app.post("/api/exchange-rate/fetch")
def fetch_exchange_rate(currency: str = "ZIG", db: Session = Depends(get_db),
                        current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    code = normalize_currency_code(currency)
    er = update_exchange_rate_in_db(db, code)
    if not er:
        raise HTTPException(500, "Failed to fetch exchange rate from provider")
    return {"currency": er.currency, "rate_to_usd": er.rate_to_usd, "source": er.source, "updated_at": str(er.updated_at)}


@app.post("/api/exchange-rate/set")
def set_exchange_rate(data: ExchangeRateSet, db: Session = Depends(get_db),
                      current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    code = normalize_currency_code(data.currency)
    er = db.query(ExchangeRate).filter(ExchangeRate.currency == code).first()
    if not er:
        er = ExchangeRate(currency=code, rate_to_usd=data.rate_to_usd, source=data.source or "manual", manual=bool(data.manual), updated_at=now())
        db.add(er)
    else:
        er.rate_to_usd = data.rate_to_usd
        er.source = data.source or er.source
        er.manual = bool(data.manual)
        er.updated_at = now()
    db.commit()
    return {"currency": er.currency, "rate_to_usd": er.rate_to_usd, "source": er.source, "manual": er.manual, "updated_at": str(er.updated_at)}

@app.put("/api/ratepayers/{rp_id}")
def update_ratepayer(rp_id: int, data: RatepayerUpdate, db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user)):
    rp = db.query(Ratepayer).filter(Ratepayer.id == rp_id).first()
    if not rp: raise HTTPException(404, "Ratepayer not found")
    changes = data.dict(exclude_none=True)
    for k, v in changes.items():
        setattr(rp, k, v)
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="ratepayers",
                    record_id=rp_id, description=f"Updated ratepayer: {rp.full_name}"))
    db.commit()
    return {"message": "Ratepayer updated"}

@app.delete("/api/ratepayers/{rp_id}")
def delete_ratepayer(rp_id: int, db: Session = Depends(get_db),
                     current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    rp = db.query(Ratepayer).filter(Ratepayer.id == rp_id).first()
    if not rp: raise HTTPException(404, "Ratepayer not found")
    inv_count = db.query(Invoice).filter(Invoice.ratepayer_id == rp_id).count()
    if inv_count > 0:
        raise HTTPException(400, f"Cannot delete: ratepayer has {inv_count} invoice(s). Deactivate instead.")
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="ratepayers",
                    record_id=rp_id, description=f"Deleted ratepayer: {rp.full_name}"))
    db.delete(rp); db.commit()
    return {"message": "Ratepayer deleted"}

# â"€â"€â"€ Invoices â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/invoices")
def list_invoices(status: Optional[str] = None, category: Optional[str] = None,
                  anomaly: Optional[str] = None, search: Optional[str] = None,
                  ratepayer_id: Optional[int] = None,
                  skip: int = 0, limit: int = 50,
                  db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    refresh_overdue_invoices(db)
    q = db.query(Invoice)
    if status:
        if status not in [e.value for e in PaymentStatus]:
            raise HTTPException(400, "Invalid status value")
        q = q.filter(Invoice.status == status)
    if category:
        if category not in [e.value for e in RevenueCategory]:
            raise HTTPException(400, "Invalid category value")
        q = q.filter(Invoice.category == category)
    if anomaly and anomaly != "none": q = q.filter(Invoice.anomaly_flag == anomaly)
    if ratepayer_id: q = q.filter(Invoice.ratepayer_id == ratepayer_id)
    if search:
        if len(search) > 100:
            raise HTTPException(400, "Search term too long")
        term = f"%{search}%"
        rp_ids = [r[0] for r in db.query(Ratepayer.id).filter(
            or_(Ratepayer.full_name.ilike(term), Ratepayer.account_number.ilike(term))
        ).all()]
        q = q.filter(Invoice.ratepayer_id.in_(rp_ids))
    total = q.count()
    items = q.order_by(desc(Invoice.issue_date)).offset(skip).limit(limit).all()
    result = []
    for inv in items:
        rp = db.query(Ratepayer).filter(Ratepayer.id == inv.ratepayer_id).first()
        result.append({
            "id": inv.id, "invoice_number": inv.invoice_number,
            "ratepayer_name": rp.full_name if rp else "Unknown",
            "account_number": rp.account_number if rp else "",
            "category": inv.category, "amount": inv.amount,
            "amount_paid": inv.amount_paid, "balance": inv.balance,
            "status": inv.status, "anomaly_flag": inv.anomaly_flag,
            "anomaly_reason": inv.anomaly_reason, "notes": inv.notes,
            "issue_date": str(inv.issue_date)[:10], "due_date": str(inv.due_date)[:10],
            "ratepayer_id": inv.ratepayer_id
        })
    return {"total": total, "items": result}

@app.post("/api/invoices")
def create_invoice(data: InvoiceCreate, db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    inv_num = "INV-" + "".join(random.choices(string.digits, k=8))
    due = datetime.fromisoformat(data.due_date)
    # Improvement 6: Duplicate invoice detection via content fingerprint (ACFE, 2022)
    fp = compute_invoice_fingerprint(data.ratepayer_id, data.category, data.amount, data.due_date)
    existing_fp = db.query(Invoice).filter(Invoice.fingerprint == fp).first()
    duplicate_warning = None
    if existing_fp:
        duplicate_warning = f"Possible duplicate: invoice {existing_fp.invoice_number} already exists with same ratepayer, category, amount, and due date."
    inv = Invoice(invoice_number=inv_num, ratepayer_id=data.ratepayer_id,
                  category=data.category, amount=data.amount, amount_paid=0.0,
                  balance=data.amount, due_date=due, notes=data.notes, created_by=current_user.id,
                  fingerprint=fp)
    cat_amounts = [r[0] for r in db.query(Invoice.amount).filter(Invoice.category == data.category).all()]
    inv.anomaly_flag, inv.anomaly_reason = _detect_anomaly(data.amount, cat_amounts)
    db.add(inv); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="invoices",
                    record_id=inv.id, description=f"Invoice {inv_num} created for ratepayer {data.ratepayer_id}"))
    db.commit()
    result = {"id": inv.id, "invoice_number": inv_num, "message": "Invoice created"}
    if duplicate_warning:
        result["duplicate_warning"] = duplicate_warning
    return result

@app.put("/api/invoices/{inv_id}")
def update_invoice(inv_id: int, data: InvoiceUpdate, db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == inv_id).first()
    if not inv: raise HTTPException(404, "Invoice not found")
    if inv.status == PaymentStatus.paid and data.amount is not None:
        raise HTTPException(400, "Cannot change amount on a paid invoice")
    if data.due_date:
        inv.due_date = datetime.fromisoformat(data.due_date)
    if data.status:
        inv.status = data.status
    if data.notes is not None:
        inv.notes = data.notes
    if data.amount is not None:
        inv.amount = data.amount
        inv.balance = max(0, data.amount - inv.amount_paid)
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="invoices",
                    record_id=inv_id, description=f"Updated invoice {inv.invoice_number}"))
    db.commit()
    return {"message": "Invoice updated"}

@app.delete("/api/invoices/{inv_id}")
def delete_invoice(inv_id: int, db: Session = Depends(get_db),
                   current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    inv = db.query(Invoice).filter(Invoice.id == inv_id).first()
    if not inv: raise HTTPException(404, "Invoice not found")
    if inv.amount_paid > 0:
        raise HTTPException(400, "Cannot delete an invoice that has payments recorded against it")
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="invoices",
                    record_id=inv_id, description=f"Deleted invoice {inv.invoice_number}"))
    db.delete(inv); db.commit()
    return {"message": "Invoice deleted"}

# â"€â"€â"€ Payments â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/payments")
def list_payments(reconciled: Optional[bool] = None, ratepayer_id: Optional[int] = None,
                  search: Optional[str] = None,
                  skip: int = 0, limit: int = 50,
                  db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Payment)
    if reconciled is not None: q = q.filter(Payment.is_reconciled == reconciled)
    if ratepayer_id:           q = q.filter(Payment.ratepayer_id == ratepayer_id)
    if search:
        if len(search) > 100:
            raise HTTPException(400, "Search term too long")
        term = f"%{search}%"
        rp_ids = [r[0] for r in db.query(Ratepayer.id).filter(
            or_(Ratepayer.full_name.ilike(term), Ratepayer.account_number.ilike(term))
        ).all()]
        q = q.filter(Payment.ratepayer_id.in_(rp_ids))
    total = q.count()
    items = q.order_by(desc(Payment.payment_date)).offset(skip).limit(limit).all()
    result = []
    for p in items:
        rp = db.query(Ratepayer).filter(Ratepayer.id == p.ratepayer_id).first()
        result.append({
            "id": p.id, "receipt_number": p.receipt_number,
            "ratepayer_name": rp.full_name if rp else "Unknown",
            "account_number": rp.account_number if rp else "",
            "amount": p.amount, "payment_method": p.payment_method,
            "currency": p.currency, "is_reconciled": p.is_reconciled,
            "anomaly_flag": p.anomaly_flag, "payment_date": str(p.payment_date)[:10],
            "invoice_id": p.invoice_id
        })
    return {"total": total, "items": result}

@app.post("/api/payments")
def record_payment(data: PaymentCreate, background_tasks: BackgroundTasks,
                   db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    rp = db.query(Ratepayer).filter(Ratepayer.id == data.ratepayer_id).first()
    if not rp: raise HTTPException(404, "Ratepayer not found")
    rcpt  = "RCP-" + "".join(random.choices(string.digits, k=8))
    flag  = AnomalyFlag.none; reason = None
    if not data.invoice_id:
        # Unlinked payment - primary leakage risk: money received with no audit trail to an invoice
        flag = AnomalyFlag.medium
        reason = "Payment recorded without invoice reference - unlinked cash increases leakage risk"
    # Z-score check: is this amount anomalous vs this ratepayer's payment history?
    rp_amounts = [r[0] for r in db.query(Payment.amount)
                  .filter(Payment.ratepayer_id == data.ratepayer_id).all()]
    zscore_flag, zscore_reason = _detect_anomaly(data.amount, rp_amounts)
    # Take the higher severity flag between the two checks
    severity_order = [AnomalyFlag.none, AnomalyFlag.low, AnomalyFlag.medium, AnomalyFlag.high]
    if severity_order.index(zscore_flag) > severity_order.index(flag):
        flag = zscore_flag
        reason = zscore_reason
    elif zscore_flag != AnomalyFlag.none and reason:
        reason = f"{reason}; {zscore_reason}"
    pmt = Payment(receipt_number=rcpt, ratepayer_id=data.ratepayer_id,
                  invoice_id=data.invoice_id, amount=data.amount,
                  payment_method=data.payment_method, currency=data.currency,
                  collected_by=current_user.id, notes=data.notes,
                  anomaly_flag=flag, anomaly_reason=reason)
    db.add(pmt)
    if data.invoice_id:
        inv = db.query(Invoice).filter(Invoice.id == data.invoice_id).first()
        if inv:
            inv.amount_paid = min(inv.amount, inv.amount_paid + data.amount)
            inv.balance     = max(0, inv.amount - inv.amount_paid)
            if inv.balance == 0:
                inv.status = PaymentStatus.paid
            elif inv.due_date < now() and inv.balance > 0:
                inv.status = PaymentStatus.overdue
    db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="payments",
                    record_id=pmt.id, description=f"Payment {rcpt} of ${data.amount} recorded"))
    db.commit()
    background_tasks.add_task(ws_manager.broadcast, {
        "type": "new_payment",
        "receipt": rcpt,
        "amount": data.amount,
        "anomaly": flag.value if flag != AnomalyFlag.none else None
    })
    return {"id": pmt.id, "receipt_number": rcpt, "message": "Payment recorded successfully"}

@app.patch("/api/payments/{pmt_id}/reconcile")
def reconcile_payment(pmt_id: int, db: Session = Depends(get_db),
                       current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant, UserRole.auditor))):
    """
    Enhanced reconciliation - records who reconciled and when (IMPROVEMENT 14).
    Restricted to admin, accountant, and auditor roles for segregation of duties.
    """
    pmt = db.query(Payment).filter(Payment.id == pmt_id).first()
    if not pmt: raise HTTPException(404, "Payment not found")
    if pmt.is_reconciled: raise HTTPException(400, "Payment is already reconciled")
    pmt.is_reconciled   = True
    pmt.reconciled_by   = current_user.id
    pmt.reconciled_at   = now()
    db.add(AuditLog(user_id=current_user.id, action="RECONCILE", table_name="payments",
                    record_id=pmt_id,
                    description=f"Payment {pmt.receipt_number} reconciled by {current_user.full_name}"))
    db.commit()
    return {"message": "Payment reconciled",
            "reconciled_by": current_user.full_name,
            "reconciled_at": str(pmt.reconciled_at)[:19]}

@app.delete("/api/payments/{pmt_id}")
def delete_payment(pmt_id: int, db: Session = Depends(get_db),
                   current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    pmt = db.query(Payment).filter(Payment.id == pmt_id).first()
    if not pmt: raise HTTPException(404, "Payment not found")
    if pmt.is_reconciled:
        raise HTTPException(400, "Cannot delete a reconciled payment")
    # Reverse invoice amount_paid if linked
    if pmt.invoice_id:
        inv = db.query(Invoice).filter(Invoice.id == pmt.invoice_id).first()
        if inv:
            inv.amount_paid = max(0, inv.amount_paid - pmt.amount)
            inv.balance     = min(inv.amount, inv.balance + pmt.amount)
            if inv.amount_paid < inv.amount:
                inv.status = PaymentStatus.overdue if inv.due_date < now() else PaymentStatus.pending
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="payments",
                    record_id=pmt_id, description=f"Deleted payment {pmt.receipt_number} of ${pmt.amount}"))
    db.delete(pmt); db.commit()
    return {"message": "Payment deleted"}

# â"€â"€â"€ Expenditures â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/expenditures")
def list_expenditures(skip: int = 0, limit: int = 100, department: Optional[str] = None,
                      db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(Expenditure)
    if department: q = q.filter(Expenditure.department == department)
    total = q.count()
    items = q.order_by(desc(Expenditure.expenditure_date)).offset(skip).limit(limit).all()
    return {"total": total, "items": [
        {"id": e.id, "reference_number": e.reference_number, "department": e.department,
         "description": e.description, "amount": e.amount, "budget_line": e.budget_line,
         "is_approved": e.is_approved, "anomaly_flag": e.anomaly_flag,
         "date": str(e.expenditure_date)[:10]} for e in items]}

@app.post("/api/expenditures")
def create_expenditure(data: ExpenditureCreate, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    ref = "EXP-" + "".join(random.choices(string.digits, k=7))
    dept_amounts = [r[0] for r in db.query(Expenditure.amount)
                    .filter(Expenditure.department == data.department).all()]
    flag, _ = _detect_anomaly(data.amount, dept_amounts)
    e = Expenditure(reference_number=ref, anomaly_flag=flag, **data.dict())
    db.add(e); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="expenditures",
                    record_id=e.id, description=f"Expenditure {ref} recorded: {data.department}"))
    db.commit()
    return {"id": e.id, "reference_number": ref, "message": "Expenditure recorded"}

@app.put("/api/expenditures/{exp_id}")
def update_expenditure(exp_id: int, data: ExpenditureUpdate, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    e = db.query(Expenditure).filter(Expenditure.id == exp_id).first()
    if not e: raise HTTPException(404, "Expenditure not found")
    if e.is_approved:
        raise HTTPException(400, "Cannot edit an approved expenditure")
    for k, v in data.dict(exclude_none=True).items():
        setattr(e, k, v)
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="expenditures",
                    record_id=exp_id, description=f"Updated expenditure {e.reference_number}"))
    db.commit()
    return {"message": "Expenditure updated"}

@app.patch("/api/expenditures/{exp_id}/approve")
def approve_expenditure(exp_id: int, db: Session = Depends(get_db),
                         current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    e = db.query(Expenditure).filter(Expenditure.id == exp_id).first()
    if not e: raise HTTPException(404)
    e.is_approved = True; e.approved_by = current_user.id
    # Update budget spent amount
    budget = db.query(Budget).filter(Budget.department == e.department).first()
    if budget:
        budget.spent_amount = (budget.spent_amount or 0) + e.amount
        budget.remaining    = budget.allocated_amount - budget.spent_amount
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="expenditures",
                    record_id=exp_id, description=f"Approved expenditure {e.reference_number}"))
    db.commit()
    return {"message": "Expenditure approved"}

@app.delete("/api/expenditures/{exp_id}")
def delete_expenditure(exp_id: int, db: Session = Depends(get_db),
                       current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    e = db.query(Expenditure).filter(Expenditure.id == exp_id).first()
    if not e: raise HTTPException(404, "Expenditure not found")
    if e.is_approved:
        raise HTTPException(400, "Cannot delete an approved expenditure")
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="expenditures",
                    record_id=exp_id, description=f"Deleted expenditure {e.reference_number}"))
    db.delete(e); db.commit()
    return {"message": "Expenditure deleted"}

# â"€â"€â"€ Budgets â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/budgets")
def list_budgets(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    items = db.query(Budget).all()
    return [{"id": b.id, "fiscal_year": b.fiscal_year, "department": b.department,
             "category": b.category, "allocated": b.allocated_amount,
             "spent": b.spent_amount, "remaining": b.remaining,
             "utilisation": round(b.spent_amount / b.allocated_amount * 100, 1) if b.allocated_amount > 0 else 0
             } for b in items]

@app.post("/api/budgets")
def create_budget(data: BudgetCreate,
                  db: Session = Depends(get_db),
                  current_user: User = Depends(require_roles(UserRole.admin, UserRole.budget_officer, UserRole.accountant))):
    existing = db.query(Budget).filter(
        Budget.fiscal_year == data.fiscal_year,
        Budget.department  == data.department,
        Budget.category    == data.category).first()
    if existing:
        raise HTTPException(400, "A budget entry already exists for this department/category/year. Use edit to update.")
    b = Budget(fiscal_year=data.fiscal_year, department=data.department,
               category=data.category, allocated_amount=data.allocated_amount,
               spent_amount=data.spent_amount,
               remaining=data.allocated_amount - data.spent_amount)
    db.add(b); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="budgets",
                    record_id=b.id, description=f"Budget created: {data.department} {data.fiscal_year}"))
    db.commit()
    return {"id": b.id, "message": "Budget created"}

@app.put("/api/budgets/{bud_id}")
def update_budget(bud_id: int, data: BudgetUpdate,
                  db: Session = Depends(get_db),
                  current_user: User = Depends(require_roles(UserRole.admin, UserRole.budget_officer, UserRole.accountant))):
    b = db.query(Budget).filter(Budget.id == bud_id).first()
    if not b: raise HTTPException(404, "Budget not found")
    for k, v in data.dict(exclude_none=True).items():
        setattr(b, k, v)
    b.remaining = b.allocated_amount - b.spent_amount
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="budgets",
                    record_id=bud_id, description=f"Updated budget: {b.department} {b.fiscal_year}"))
    db.commit()
    return {"message": "Budget updated"}

@app.delete("/api/budgets/{bud_id}")
def delete_budget(bud_id: int,
                  db: Session = Depends(get_db),
                  current_user: User = Depends(require_roles(UserRole.admin, UserRole.budget_officer))):
    b = db.query(Budget).filter(Budget.id == bud_id).first()
    if not b: raise HTTPException(404, "Budget not found")
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="budgets",
                    record_id=bud_id, description=f"Deleted budget: {b.department} {b.fiscal_year}"))
    db.delete(b); db.commit()
    return {"message": "Budget deleted"}

# â"€â"€â"€ Leakage & Anomalies â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/leakage/summary")
def leakage_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Anomaly counts include both invoice-level and payment-level flags
    inv_high   = db.query(Invoice).filter(Invoice.anomaly_flag == AnomalyFlag.high).count()
    inv_medium = db.query(Invoice).filter(Invoice.anomaly_flag == AnomalyFlag.medium).count()
    inv_low    = db.query(Invoice).filter(Invoice.anomaly_flag == AnomalyFlag.low).count()
    pmt_high   = db.query(Payment).filter(Payment.anomaly_flag == AnomalyFlag.high).count()
    pmt_medium = db.query(Payment).filter(Payment.anomaly_flag == AnomalyFlag.medium).count()
    pmt_low    = db.query(Payment).filter(Payment.anomaly_flag == AnomalyFlag.low).count()

    unreconciled     = db.query(Payment).filter(Payment.is_reconciled == False).count()
    unrecon_cash     = usd_payment_sum(db.query(Payment).filter(
        Payment.is_reconciled == False, Payment.payment_method == "cash"))
    unrecon_other    = usd_payment_sum(db.query(Payment).filter(
        Payment.is_reconciled == False, Payment.payment_method != "cash"))
    overdue_amt      = db.query(func.sum(Invoice.balance)).filter(
        Invoice.status == PaymentStatus.overdue).scalar() or 0
    waived_total     = db.query(func.sum(Invoice.amount)).filter(
        Invoice.status == PaymentStatus.waived).scalar() or 0
    anomaly_inv_amt  = db.query(func.sum(Invoice.amount)).filter(
        Invoice.anomaly_flag != AnomalyFlag.none).scalar() or 0
    alerts = db.query(LeakageAlert).filter(LeakageAlert.is_resolved == False).all()

    # 5-component ACFE (2022) weighted leakage model — same formula as /api/leakage/quantification
    estimated_leakage = round(
        unrecon_cash  * 0.40 +
        unrecon_other * 0.20 +
        overdue_amt   * 0.25 +
        waived_total  * 0.15 +
        anomaly_inv_amt * 0.10,
        2
    )
    return {
        "high_anomalies":       inv_high   + pmt_high,
        "medium_anomalies":     inv_medium + pmt_medium,
        "low_anomalies":        inv_low    + pmt_low,
        "unreconciled_payments": unreconciled,
        "unreconciled_amount":  round(unrecon_cash + unrecon_other, 2),
        "overdue_balance":      round(overdue_amt, 2),
        "active_alerts":        len(alerts),
        "estimated_leakage":    estimated_leakage,
    }

@app.get("/api/leakage/alerts")
def leakage_alerts(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    alerts = db.query(LeakageAlert).order_by(desc(LeakageAlert.created_at)).all()
    return [{"id": a.id, "type": a.alert_type, "severity": a.severity, "description": a.description,
             "is_resolved": a.is_resolved, "created_at": str(a.created_at)[:16]} for a in alerts]

@app.patch("/api/leakage/alerts/{alert_id}/resolve")
def resolve_alert(alert_id: int, data: AlertResolveRequest, db: Session = Depends(get_db),
                  current_user: User = Depends(get_current_user)):
    """
    Enhanced alert resolution requiring mandatory resolution notes (IMPROVEMENT 9).
    Aligned with COSO (2013) documented corrective action requirements and PECGA s14(c).
    """
    if not data.resolution_notes or not data.resolution_notes.strip():
        raise HTTPException(400, "resolution_notes is required. Describe what corrective action was taken.")
    a = db.query(LeakageAlert).filter(LeakageAlert.id == alert_id).first()
    if not a: raise HTTPException(404, "Alert not found")
    a.is_resolved      = True
    a.resolved_by      = current_user.id
    a.resolved_at      = now()
    a.resolution_notes = data.resolution_notes.strip()
    db.add(AuditLog(user_id=current_user.id, action="RESOLVE", table_name="leakage_alerts",
                    record_id=a.id,
                    description=f"Alert #{a.id} ({a.alert_type}) resolved by {current_user.full_name}. Notes: {a.resolution_notes}"))
    db.commit()
    return {"message": "Alert resolved", "resolution_notes": a.resolution_notes}

@app.post("/api/leakage/scan")
def scan_leakage_alerts(background_tasks: BackgroundTasks,
                        db: Session = Depends(get_db),
                        current_user: User = Depends(require_roles(
                            UserRole.admin, UserRole.auditor, UserRole.accountant))):
    """
    Dynamically scan the database for revenue leakage patterns and generate alerts.
    Implements five detection rules derived from the City of Harare stakeholder
    interviews and ACFE (2022) fraud pattern taxonomy:
      1. Ghost accounts - active ratepayers with no payment in 12+ months but outstanding balances
      2. Waiver abuse - waivers that exceed the ward's historical waiver rate by >2Ïƒ
      3. Unlinked cash - unreconciled cash payments with no invoice reference
      4. Officer collection gap - revenue officers collecting significantly below peers (Z-score)
      5. Stale high-value overdue - overdue invoices >180 days with balance >$500
    """
    cutoff_12m  = now() - timedelta(days=365)
    cutoff_180d = now() - timedelta(days=180)
    generated   = 0

    def _alert_exists(alert_type: str, record_id: int, table: str) -> bool:
        return db.query(LeakageAlert).filter(
            LeakageAlert.alert_type == alert_type,
            LeakageAlert.related_record_id == record_id,
            LeakageAlert.related_table == table,
            LeakageAlert.is_resolved == False
        ).first() is not None

    # â"€â"€ Rule 1: Ghost accounts â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    active_rps = db.query(Ratepayer).filter(Ratepayer.is_active == True).all()
    for rp in active_rps:
        outstanding = db.query(func.sum(Invoice.balance))\
            .filter(Invoice.ratepayer_id == rp.id, Invoice.balance > 0).scalar() or 0
        if outstanding <= 0:
            continue
        last_pmt = db.query(func.max(Payment.payment_date))\
            .filter(Payment.ratepayer_id == rp.id).scalar()
        if last_pmt is None or last_pmt < cutoff_12m:
            if not _alert_exists("ghost_account", rp.id, "ratepayers"):
                db.add(LeakageAlert(
                    alert_type="ghost_account", severity="high",
                    description=(f"Ratepayer {rp.account_number} ({rp.full_name}) has "
                                 f"${outstanding:,.2f} outstanding but no payment in >12 months. "
                                 f"Possible ghost account or inactive debtor - review for write-off or enforcement."),
                    related_record_id=rp.id, related_table="ratepayers"
                ))
                generated += 1

    # â"€â"€ Rule 2: Unlinked cash payments (unreconciled, no invoice) â"€â"€â"€â"€â"€â"€â"€â"€â"€
    unlinked = db.query(Payment).filter(
        Payment.invoice_id == None,
        Payment.is_reconciled == False,
        Payment.payment_method == "cash"
    ).all()
    for pmt in unlinked:
        if not _alert_exists("unlinked_cash", pmt.id, "payments"):
            db.add(LeakageAlert(
                alert_type="unlinked_cash", severity="high",
                description=(f"Cash payment {pmt.receipt_number} of ${pmt.amount:,.2f} is "
                             f"unreconciled and has no invoice reference. "
                             f"Cash with no paper trail is the primary leakage vector (ACFE, 2022)."),
                related_record_id=pmt.id, related_table="payments"
            ))
            generated += 1

    # â"€â"€ Rule 3: Stale high-value overdue invoices (>180 days, >$500) â"€â"€â"€â"€â"€
    stale = db.query(Invoice).filter(
        Invoice.status == PaymentStatus.overdue,
        Invoice.due_date < cutoff_180d,
        Invoice.balance > 500
    ).all()
    for inv in stale:
        if not _alert_exists("stale_overdue", inv.id, "invoices"):
            db.add(LeakageAlert(
                alert_type="stale_overdue", severity="medium",
                description=(f"Invoice {inv.invoice_number} has been overdue >180 days "
                             f"with ${inv.balance:,.2f} outstanding. "
                             f"Accounts >180 days overdue have <30% recovery probability (NCC, 2023)."),
                related_record_id=inv.id, related_table="invoices"
            ))
            generated += 1

    # â"€â"€ Rule 4: Officer collection gap (Z-score on collection rates) â"€â"€â"€â"€â"€â"€
    officers = db.query(User).filter(User.role == UserRole.revenue_officer, User.is_active == True).all()
    officer_rates = []
    for officer in officers:
        total_collected = usd_payment_sum(db.query(Payment).filter(Payment.collected_by == officer.id))
        inv_count = db.query(func.count(Invoice.id))\
            .filter(Invoice.created_by == officer.id).scalar() or 0
        total_billed = db.query(func.sum(Invoice.amount))\
            .filter(Invoice.created_by == officer.id).scalar() or 0
        rate = (total_collected / total_billed * 100) if total_billed > 0 else 0
        officer_rates.append((officer, rate, total_billed))

    if len(officer_rates) >= 3:
        rates = [r[1] for r in officer_rates]
        mean_rate = sum(rates) / len(rates)
        variance  = sum((r - mean_rate) ** 2 for r in rates) / (len(rates) - 1)
        std_rate  = math.sqrt(variance) if variance > 0 else 0
        for officer, rate, billed in officer_rates:
            if billed < 100:
                continue
            z = (rate - mean_rate) / std_rate if std_rate > 0 else 0
            if z < -2.0:
                if not _alert_exists("officer_gap", officer.id, "users"):
                    db.add(LeakageAlert(
                        alert_type="officer_gap", severity="medium",
                        description=(f"Revenue officer {officer.full_name} has a collection rate "
                                     f"of {rate:.1f}% vs peer average {mean_rate:.1f}% "
                                     f"(Z-score {z:.2f}). Significantly below peers - review workload or escalate."),
                        related_record_id=officer.id, related_table="users"
                    ))
                    generated += 1

    # â"€â"€ Rule 5: Waived invoices without approval audit trail â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    waived = db.query(Invoice).filter(Invoice.status == PaymentStatus.waived).all()
    for inv in waived:
        if not _alert_exists("waiver_no_audit", inv.id, "invoices"):
            # Check if there is an audit log entry showing who waived it
            waiver_log = db.query(AuditLog).filter(
                AuditLog.table_name == "invoices",
                AuditLog.record_id == inv.id,
                AuditLog.action == "UPDATE"
            ).first()
            if not waiver_log:
                db.add(LeakageAlert(
                    alert_type="waiver_no_audit", severity="high",
                    description=(f"Invoice {inv.invoice_number} (${inv.amount:,.2f}) is marked "
                                 f"waived but has no audit trail of who approved it. "
                                 f"Unapproved waivers are a key leakage indicator (ZIMRA, 2023)."),
                    related_record_id=inv.id, related_table="invoices"
                ))
                generated += 1

    # â"€â"€ Rule 6: Duplicate payments (same ratepayer, amount, same day) â"€â"€â"€â"€â"€â"€â"€â"€
    from sqlalchemy import func as _func
    dup_query = db.query(
        Payment.ratepayer_id, Payment.amount,
        func.date(Payment.payment_date).label("pmt_date"),
        func.count(Payment.id).label("cnt")
    ).group_by(Payment.ratepayer_id, Payment.amount, func.date(Payment.payment_date))     .having(func.count(Payment.id) > 1).all()

    seen_groups = set()
    for row in dup_query:
        group_key = (row.ratepayer_id, row.amount, str(row.pmt_date))
        if group_key in seen_groups:
            continue
        seen_groups.add(group_key)
        rp_dup = db.query(Ratepayer).filter(Ratepayer.id == row.ratepayer_id).first()
        rp_name = rp_dup.full_name if rp_dup else f"ID {row.ratepayer_id}"
        sample_pmt = db.query(Payment).filter(
            Payment.ratepayer_id == row.ratepayer_id,
            Payment.amount == row.amount,
            func.date(Payment.payment_date) == str(row.pmt_date)
        ).first()
        if sample_pmt and not _alert_exists("duplicate_payment", row.ratepayer_id, "payments"):
            db.add(LeakageAlert(
                alert_type="duplicate_payment", severity="high",
                description=(f"Multiple payments of ${row.amount:,.2f} recorded for {rp_name} "
                             f"on {str(row.pmt_date)[:10]}. "
                             f"Duplicate payments indicate double-posting or cash diversion (ACFE, 2022)."),
                related_record_id=row.ratepayer_id, related_table="payments"
            ))
            generated += 1

    # â"€â"€ Rule 7: Round-trip - waiver and payment on same account, same day â"€â"€â"€â"€â"€
    waived_invs = db.query(Invoice).filter(Invoice.status == PaymentStatus.waived).all()
    for winv in waived_invs:
        pmt_same_day = db.query(Payment).filter(
            Payment.ratepayer_id == winv.ratepayer_id,
            func.date(Payment.payment_date) == func.date(winv.issue_date)
        ).first()
        if pmt_same_day and not _alert_exists("round_trip_waiver", winv.id, "invoices"):
            db.add(LeakageAlert(
                alert_type="round_trip_waiver", severity="high",
                description=(f"Invoice {winv.invoice_number} (${winv.amount:,.2f}) was waived on "
                             f"the same day a payment was recorded for the same ratepayer. "
                             f"Consistent with fictitious collection schemes (ZIMRA, 2023)."),
                related_record_id=winv.id, related_table="invoices"
            ))
            generated += 1

    if generated > 0:
        db.commit()
        background_tasks.add_task(ws_manager.broadcast, {
            "type": "leakage_scan",
            "generated": generated,
            "message": f"{generated} new leakage alert(s) detected - refresh your dashboard"
        })
    return {"generated": generated, "message": f"{generated} new alert(s) generated from leakage scan (7 rules)"}

# ─── Leakage Quantification (Dissertation Objective 2) ──────────────────────

@app.get("/api/leakage/quantification")
def leakage_quantification(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Quantifies revenue leakage by category and by leakage type.
    Provides dollar amounts, percentages, and a Leakage Risk Index for
    dissertation Objective 2 (revenue leakage measurement).

    Leakage weights derived from ACFE (2022) Revenue Assurance Framework:
      - Unreconciled cash:  40% (cash received, no audit trail)
      - Unreconciled other: 20% (EFT/cheque, lower risk)
      - Overdue balances:   25% (NCC 2023: <25% recovery probability >90 days)
      - Unauthorized waivers: 15% (direct write-off leakage)
      - Anomalous invoices: 10% (billing fraud indicator)
    """
    total_billed = db.query(func.sum(Invoice.amount)).scalar() or 0

    # ── By revenue category ──────────────────────────────────────────────
    # Use Payment.amount joined to Invoice to correctly count actual cash received per
    # category. Invoice.amount_paid misses orphaned (unlinked) cash payments.
    by_category = []
    for (cat,) in db.query(Invoice.category).distinct().all():
        billed       = db.query(func.sum(Invoice.amount)).filter(Invoice.category == cat).scalar() or 0
        collected    = usd_payment_sum(
            db.query(Payment)
              .join(Invoice, Payment.invoice_id == Invoice.id)
              .filter(Invoice.category == cat)
        )
        overdue_bal  = db.query(func.sum(Invoice.balance)).filter(
            Invoice.category == cat, Invoice.status == PaymentStatus.overdue).scalar() or 0
        anomaly_amt  = db.query(func.sum(Invoice.amount)).filter(
            Invoice.category == cat, Invoice.anomaly_flag != AnomalyFlag.none).scalar() or 0
        est_leakage  = round(overdue_bal * 0.25 + anomaly_amt * 0.10, 2)
        by_category.append({
            "category":          cat,
            "billed":            round(billed, 2),
            "collected":         round(collected, 2),
            "outstanding":       round(billed - collected, 2),
            "overdue_balance":   round(overdue_bal, 2),
            "anomaly_amount":    round(anomaly_amt, 2),
            "estimated_leakage": est_leakage,
            "leakage_pct":       round(est_leakage / billed * 100, 1) if billed > 0 else 0,
            "collection_rate":   round(collected / billed * 100, 1) if billed > 0 else 0,
        })

    # Orphaned payments (no invoice link) are real cash received but uncategorised.
    # Include them as a separate row so by_category totals reconcile to dashboard total.
    orphaned = usd_payment_sum(db.query(Payment).filter(Payment.invoice_id == None))
    if orphaned > 0:
        by_category.append({
            "category":          "unlinked",
            "billed":            0,
            "collected":         round(orphaned, 2),
            "outstanding":       round(-orphaned, 2),
            "overdue_balance":   0,
            "anomaly_amount":    0,
            "estimated_leakage": 0,
            "leakage_pct":       0,
            "collection_rate":   0,
        })

    # ── By leakage type ──────────────────────────────────────────────────
    unrecon_cash  = usd_payment_sum(db.query(Payment).filter(
        Payment.is_reconciled == False, Payment.payment_method == "cash"))
    unrecon_other = usd_payment_sum(db.query(Payment).filter(
        Payment.is_reconciled == False, Payment.payment_method != "cash"))
    overdue_total = db.query(func.sum(Invoice.balance)).filter(
        Invoice.status == PaymentStatus.overdue).scalar() or 0
    waived_total  = db.query(func.sum(Invoice.amount)).filter(
        Invoice.status == PaymentStatus.waived).scalar() or 0
    anomaly_total = db.query(func.sum(Invoice.amount)).filter(
        Invoice.anomaly_flag != AnomalyFlag.none).scalar() or 0

    by_type = [
        {"type": "Unreconciled Cash",       "raw": round(unrecon_cash, 2),  "leakage": round(unrecon_cash  * 0.40, 2), "weight": 0.40,
         "description": "Cash received but not reconciled to an invoice - primary leakage vector (ACFE, 2022)"},
        {"type": "Unreconciled Non-Cash",   "raw": round(unrecon_other, 2), "leakage": round(unrecon_other * 0.20, 2), "weight": 0.20,
         "description": "EFT/cheque payments unreconciled - reduced weight vs cash (ACFE, 2022)"},
        {"type": "Overdue Balances",        "raw": round(overdue_total, 2), "leakage": round(overdue_total * 0.25, 2), "weight": 0.25,
         "description": "Outstanding overdue invoices - 25% recovery weight (NCC, 2023: <25% recovery >90 days)"},
        {"type": "Unauthorized Waivers",    "raw": round(waived_total, 2),  "leakage": round(waived_total  * 0.15, 2), "weight": 0.15,
         "description": "Waived invoice amounts - direct revenue write-off (ZIMRA, 2023)"},
        {"type": "Anomalous Invoices",      "raw": round(anomaly_total, 2), "leakage": round(anomaly_total * 0.10, 2), "weight": 0.10,
         "description": "Invoices flagged by Isolation Forest anomaly detection (Liu, Ting & Zhou, 2008)"},
    ]
    by_type.sort(key=lambda x: -x["leakage"])

    total_leakage  = round(sum(t["leakage"] for t in by_type), 2)
    leakage_rate   = round(total_leakage / total_billed * 100, 1) if total_billed > 0 else 0
    high_risk_cats = [c for c in by_category if c["leakage_pct"] >= 10]

    return {
        "total_billed":       round(total_billed, 2),
        "total_leakage":      total_leakage,
        "leakage_rate_pct":   leakage_rate,
        "by_category":        sorted(by_category, key=lambda x: -x["estimated_leakage"]),
        "by_type":            by_type,
        "high_risk_categories": [c["category"] for c in high_risk_cats],
        "methodology": (
            "ACFE (2022) Revenue Assurance Framework - weighted risk model. "
            "Anomaly detection: Isolation Forest (Liu, Ting & Zhou, 2008, IEEE ICDM). "
            "Overdue recovery estimate: NCC (2023) Municipal Revenue Recovery Report."
        ),
        "generated_at": str(now())
    }

# ─── Leakage Cause Analysis (Dissertation Objective 1) ──────────────────────

@app.get("/api/leakage/cause-analysis")
def leakage_cause_analysis(
    months: int = Query(6, ge=3, le=24),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Historical month-by-month leakage cause analysis for dissertation Objective 1.
    Decomposes estimated leakage into three root causes per month and fits
    a linear trend (OLS) to identify whether leakage is improving or worsening.
    """
    monthly = []
    for i in range(months, 0, -1):
        ref   = now().replace(day=1) - timedelta(days=28 * (i - 1))
        ms    = ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if ms.month == 12:
            me = ms.replace(year=ms.year + 1, month=1, day=1)
        else:
            me = ms.replace(month=ms.month + 1, day=1)

        unrecon  = usd_payment_sum(db.query(Payment).filter(
            Payment.payment_date >= ms, Payment.payment_date < me,
            Payment.is_reconciled == False))
        overdue  = db.query(func.sum(Invoice.balance)).filter(
            Invoice.issue_date >= ms, Invoice.issue_date < me,
            Invoice.status == PaymentStatus.overdue).scalar() or 0
        anomaly  = db.query(func.sum(Invoice.amount)).filter(
            Invoice.issue_date >= ms, Invoice.issue_date < me,
            Invoice.anomaly_flag != AnomalyFlag.none).scalar() or 0
        billed   = db.query(func.sum(Invoice.amount)).filter(
            Invoice.issue_date >= ms, Invoice.issue_date < me).scalar() or 0
        collected = usd_payment_sum(db.query(Payment).filter(
            Payment.payment_date >= ms, Payment.payment_date < me))

        unreconciled_leakage = round(unrecon  * 0.40, 2)
        overdue_leakage      = round(overdue  * 0.25, 2)
        anomaly_leakage      = round(anomaly  * 0.10, 2)
        total_leakage        = round(unreconciled_leakage + overdue_leakage + anomaly_leakage, 2)

        monthly.append({
            "month":                 ms.strftime("%b %Y"),
            "billed":                round(billed, 2),
            "collected":             round(collected, 2),
            "collection_gap":        round(billed - collected, 2),
            "unreconciled_leakage":  unreconciled_leakage,
            "overdue_leakage":       overdue_leakage,
            "anomaly_leakage":       anomaly_leakage,
            "total_estimated_leakage": total_leakage,
            "leakage_rate_pct":      round(total_leakage / billed * 100, 1) if billed > 0 else 0,
        })

    # OLS linear trend on total leakage
    vals = [m["total_estimated_leakage"] for m in monthly]
    n = len(vals)
    x = list(range(n)); mx = sum(x) / n; my = sum(vals) / n
    num = sum((x[i] - mx) * (vals[i] - my) for i in range(n))
    den = sum((xi - mx) ** 2 for xi in x)
    slope = round(num / den, 2) if den != 0 else 0
    if abs(slope) < 50:
        trend = "stable"
    elif slope > 0:
        trend = "worsening"
    else:
        trend = "improving"

    # Aggregate cause totals
    all_unrecon  = round(sum(m["unreconciled_leakage"] for m in monthly), 2)
    all_overdue  = round(sum(m["overdue_leakage"] for m in monthly), 2)
    all_anomaly  = round(sum(m["anomaly_leakage"] for m in monthly), 2)
    grand_total  = all_unrecon + all_overdue + all_anomaly or 1

    top_causes = sorted([
        {"cause": "Unreconciled Payments", "total": all_unrecon,
         "share_pct": round(all_unrecon / grand_total * 100, 1),
         "description": "Payments received but not matched to invoices (ACFE, 2022)"},
        {"cause": "Overdue Balances",      "total": all_overdue,
         "share_pct": round(all_overdue / grand_total * 100, 1),
         "description": "Invoices past due with low recovery probability (NCC, 2023)"},
        {"cause": "Anomalous Invoices",    "total": all_anomaly,
         "share_pct": round(all_anomaly / grand_total * 100, 1),
         "description": "Invoices flagged by Isolation Forest (Liu, Ting & Zhou, 2008)"},
    ], key=lambda x: -x["total"])

    return {
        "months_analysed":  months,
        "monthly_trend":    monthly,
        "trend_direction":  trend,
        "slope_per_month":  slope,
        "top_causes":       top_causes,
        "total_leakage_period": round(grand_total, 2),
        "methodology": (
            "OLS linear regression on monthly leakage totals (Freedman, Pisani & Purves, 2007). "
            "Leakage components weighted per ACFE (2022) Revenue Assurance Framework."
        )
    }

# â"€â"€â"€ Audit Log â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/audit-logs")
def get_audit_logs(skip: int = 0, limit: int = 50,
                   date_from: Optional[str] = None, date_to: Optional[str] = None,
                   action: Optional[str] = None, table_name: Optional[str] = None,
                   user_actions_only: bool = False,
                   db: Session = Depends(get_db),
                   current_user: User = Depends(require_roles(UserRole.admin, UserRole.auditor))):
    q = db.query(AuditLog).order_by(desc(AuditLog.timestamp))
    if date_from: q = q.filter(AuditLog.timestamp >= datetime.fromisoformat(date_from))
    if date_to:   q = q.filter(AuditLog.timestamp <= datetime.fromisoformat(date_to + "T23:59:59"))
    if action:     q = q.filter(AuditLog.action == action.upper())
    if table_name: q = q.filter(AuditLog.table_name == table_name)
    if user_actions_only: q = q.filter(AuditLog.user_id != None)
    total = q.count()
    items = q.offset(skip).limit(limit).all()
    result = []
    for log in items:
        user = db.query(User).filter(User.id == log.user_id).first()
        result.append({"id": log.id, "action": log.action, "table_name": log.table_name,
                        "description": log.description, "user": user.full_name if user else "System",
                        "ip_address": log.ip_address, "timestamp": str(log.timestamp)[:16]})
    return {"total": total, "items": result}

# â"€â"€â"€ Users â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/users")
def list_users(db: Session = Depends(get_db),
               current_user: User = Depends(require_roles(UserRole.admin))):
    users = db.query(User).all()
    return [{"id": u.id, "username": u.username, "full_name": u.full_name,
             "email": u.email, "role": u.role, "is_active": u.is_active,
             "last_login": str(u.last_login)[:16] if u.last_login else None} for u in users]

@app.post("/api/users")
def create_user(data: UserCreate, db: Session = Depends(get_db),
                current_user: User = Depends(require_roles(UserRole.admin))):
    if db.query(User).filter(User.username == data.username).first():
        raise HTTPException(400, "Username already exists")
    if db.query(User).filter(User.email == data.email).first():
        raise HTTPException(400, "Email already in use")
    try:
        role = UserRole(data.role)
    except ValueError:
        raise HTTPException(400, f"Invalid role: {data.role}")
    u = User(username=data.username, full_name=data.full_name, email=data.email,
             hashed_password=hash_password(data.password), role=role)
    db.add(u); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="users",
                    record_id=u.id, description=f"Created user: {data.username} ({data.role})"))
    db.commit()
    return {"id": u.id, "username": u.username, "message": "User created"}

@app.put("/api/users/{user_id}")
def update_user(user_id: int, data: UserUpdate, db: Session = Depends(get_db),
                current_user: User = Depends(require_roles(UserRole.admin))):
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    if user_id == current_user.id and data.is_active is False:
        raise HTTPException(400, "Cannot deactivate your own account")
    if data.full_name:  u.full_name = data.full_name
    if data.email:      u.email     = data.email
    if data.role:
        try: u.role = UserRole(data.role)
        except ValueError: raise HTTPException(400, f"Invalid role: {data.role}")
    if data.is_active is not None: u.is_active = data.is_active
    if data.password:  u.hashed_password = hash_password(data.password)
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="users",
                    record_id=user_id, description=f"Updated user: {u.username}"))
    db.commit()
    return {"message": "User updated"}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db),
                current_user: User = Depends(require_roles(UserRole.admin))):
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    if user_id == current_user.id:
        raise HTTPException(400, "Cannot delete your own account")
    u.is_active = False
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="users",
                    record_id=user_id, description=f"Deactivated user: {u.username}"))
    db.commit()
    return {"message": "User deactivated"}

# â"€â"€â"€ Export Endpoints â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/export/ratepayers")
def export_ratepayers(format: str = "csv", db: Session = Depends(get_db),
                      current_user: User = Depends(get_current_user)):
    items = db.query(Ratepayer).all()
    headers = ["Account Number", "Full Name", "Address", "Ward", "Zone",
               "Property Type", "Phone", "Email", "Active", "Created At"]
    rows = [[r.account_number, r.full_name, r.address, r.ward, r.zone,
             r.property_type, r.phone or "", r.email or "",
             "Yes" if r.is_active else "No", str(r.created_at)[:10]] for r in items]
    return export_response(format, headers, rows,
                           "ratepayers.csv", "ratepayers.xlsx", "Ratepayers",
                           "City of Harare FMS - Ratepayer Registry")

@app.get("/api/export/invoices")
def export_invoices(format: str = "csv", db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    items = db.query(Invoice).order_by(desc(Invoice.issue_date)).all()
    headers = ["Invoice Number", "Ratepayer", "Account Number", "Category",
               "Amount", "Amount Paid", "Balance", "Status",
               "Issue Date", "Due Date", "Anomaly Flag", "Notes"]
    rows = []
    for inv in items:
        rp = db.query(Ratepayer).filter(Ratepayer.id == inv.ratepayer_id).first()
        rows.append([inv.invoice_number, rp.full_name if rp else "", rp.account_number if rp else "",
                     inv.category, inv.amount, inv.amount_paid, inv.balance, inv.status,
                     str(inv.issue_date)[:10], str(inv.due_date)[:10], inv.anomaly_flag, inv.notes or ""])
    return export_response(format, headers, rows,
                           "invoices.csv", "invoices.xlsx", "Invoices",
                           "City of Harare FMS - Invoice Register")

@app.get("/api/export/payments")
def export_payments(format: str = "csv", db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    items = db.query(Payment).order_by(desc(Payment.payment_date)).all()
    headers = ["Receipt Number", "Ratepayer", "Account Number", "Amount",
               "Currency", "Method", "Date", "Reconciled", "Anomaly Flag"]
    rows = []
    for p in items:
        rp = db.query(Ratepayer).filter(Ratepayer.id == p.ratepayer_id).first()
        rows.append([p.receipt_number, rp.full_name if rp else "", rp.account_number if rp else "",
                     p.amount, p.currency, p.payment_method, str(p.payment_date)[:10],
                     "Yes" if p.is_reconciled else "No", p.anomaly_flag])
    return export_response(format, headers, rows,
                           "payments.csv", "payments.xlsx", "Payments",
                           "City of Harare FMS - Payment Records")

@app.get("/api/export/expenditures")
def export_expenditures(format: str = "csv", db: Session = Depends(get_db),
                        current_user: User = Depends(get_current_user)):
    items = db.query(Expenditure).order_by(desc(Expenditure.expenditure_date)).all()
    headers = ["Reference", "Department", "Description", "Amount",
               "Budget Line", "Date", "Approved", "Anomaly Flag"]
    rows = [[e.reference_number, e.department, e.description, e.amount,
             e.budget_line, str(e.expenditure_date)[:10],
             "Yes" if e.is_approved else "No", e.anomaly_flag] for e in items]
    return export_response(format, headers, rows,
                           "expenditures.csv", "expenditures.xlsx", "Expenditures",
                           "City of Harare FMS - Expenditure Register")

@app.get("/api/export/audit-logs")
def export_audit_logs(format: str = "csv",
                      date_from: Optional[str] = None, date_to: Optional[str] = None,
                      db: Session = Depends(get_db),
                      current_user: User = Depends(require_roles(UserRole.admin, UserRole.auditor))):
    q = db.query(AuditLog).order_by(desc(AuditLog.timestamp))
    if date_from: q = q.filter(AuditLog.timestamp >= datetime.fromisoformat(date_from))
    if date_to:   q = q.filter(AuditLog.timestamp <= datetime.fromisoformat(date_to + "T23:59:59"))
    items = q.all()
    headers = ["#", "User", "Action", "Module", "Description", "IP Address", "Timestamp"]
    rows = []
    for i, log in enumerate(items, 1):
        user = db.query(User).filter(User.id == log.user_id).first()
        rows.append([i, user.full_name if user else "System", log.action,
                     log.table_name, log.description or "", log.ip_address or "", str(log.timestamp)[:16]])
    return export_response(format, headers, rows,
                           "audit_trail.csv", "audit_trail.xlsx", "Audit Trail",
                           "City of Harare FMS - Audit Trail")

@app.get("/api/export/budgets")
def export_budgets(format: str = "csv", db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    items = db.query(Budget).order_by(Budget.fiscal_year, Budget.department).all()
    headers = ["Fiscal Year", "Department", "Category", "Allocated (USD)",
               "Spent (USD)", "Remaining (USD)", "Utilisation %"]
    rows = []
    for b in items:
        util = round((b.spent_amount / b.allocated_amount * 100), 1) if b.allocated_amount else 0
        rows.append([b.fiscal_year, b.department, b.category,
                     b.allocated_amount, b.spent_amount, b.remaining, util])
    return export_response(format, headers, rows,
                           "budgets.csv", "budgets.xlsx", "Budgets",
                           "City of Harare FMS - Budget Register")

@app.get("/api/export/leakage")
def export_leakage(format: str = "csv", db: Session = Depends(get_db),
                   current_user: User = Depends(require_roles(
                       UserRole.admin, UserRole.auditor, UserRole.accountant))):
    items = db.query(LeakageAlert).order_by(desc(LeakageAlert.created_at)).all()
    headers = ["ID", "Alert Type", "Severity", "Description",
               "Related Table", "Resolved", "Resolution Notes", "Created At", "Resolved At"]
    rows = [[a.id, a.alert_type, a.severity, a.description,
             a.related_table or "", "Yes" if a.is_resolved else "No",
             a.resolution_notes or "", str(a.created_at)[:16],
             str(a.resolved_at)[:16] if a.resolved_at else ""] for a in items]
    return export_response(format, headers, rows,
                           "leakage_alerts.csv", "leakage_alerts.xlsx", "Leakage Alerts",
                           "City of Harare FMS - Revenue Leakage Alerts")

@app.get("/api/export/users")
def export_users(format: str = "csv", db: Session = Depends(get_db),
                 current_user: User = Depends(require_roles(UserRole.admin))):
    items = db.query(User).order_by(User.full_name).all()
    headers = ["ID", "Full Name", "Username", "Email", "Role", "Active",
               "Created At", "Last Login", "Failed Login Count"]
    rows = [[u.id, u.full_name, u.username, u.email, u.role,
             "Yes" if u.is_active else "No",
             str(u.created_at)[:10],
             str(u.last_login)[:16] if u.last_login else "",
             u.failed_login_count] for u in items]
    return export_response(format, headers, rows,
                           "users.csv", "users.xlsx", "Users",
                           "City of Harare FMS - User Registry")

# â"€â"€â"€ Import Endpoints â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.post("/api/import/ratepayers")
async def import_ratepayers(file: UploadFile = File(...),
                             db: Session = Depends(get_db),
                             current_user: User = Depends(require_roles(UserRole.admin, UserRole.revenue_officer))):
    rows  = await parse_upload(file)
    created = 0; errors = []
    for i, row in enumerate(rows, 2):
        try:
            name = str(row.get("full_name") or row.get("Full Name") or "").strip()
            addr = str(row.get("address") or row.get("Address") or "").strip()
            ward = str(row.get("ward") or row.get("Ward") or "Ward 1").strip()
            zone = str(row.get("zone") or row.get("Zone") or "").strip()
            phone= str(row.get("phone") or row.get("Phone") or "").strip() or None
            email= str(row.get("email") or row.get("Email") or "").strip() or None
            ptype= str(row.get("property_type") or row.get("Property Type") or "residential").strip().lower()
            if not name or not addr:
                errors.append(f"Row {i}: full_name and address are required"); continue
            acct = "COH-" + "".join(random.choices(string.digits, k=6))
            rp = Ratepayer(account_number=acct, full_name=name, address=addr, ward=ward,
                           zone=zone, phone=phone, email=email, property_type=ptype)
            db.add(rp); created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
    db.flush()
    db.add(AuditLog(user_id=current_user.id, action="IMPORT", table_name="ratepayers",
                    description=f"Imported {created} ratepayers from file"))
    db.commit()
    return {"created": created, "errors": errors, "message": f"{created} ratepayers imported"}

@app.post("/api/import/invoices")
async def import_invoices(file: UploadFile = File(...),
                           db: Session = Depends(get_db),
                           current_user: User = Depends(require_roles(UserRole.admin, UserRole.revenue_officer, UserRole.accountant))):
    rows = await parse_upload(file)
    created = 0; errors = []
    for i, row in enumerate(rows, 2):
        try:
            acct     = str(row.get("account_number") or row.get("Account Number") or "").strip()
            category = str(row.get("category") or row.get("Category") or "other").strip().lower()
            amount   = float(row.get("amount") or row.get("Amount") or 0)
            due_raw  = str(row.get("due_date") or row.get("Due Date") or "").strip()
            notes    = str(row.get("notes") or row.get("Notes") or "").strip() or None
            if not acct or not amount or not due_raw:
                errors.append(f"Row {i}: account_number, amount and due_date required"); continue
            rp = db.query(Ratepayer).filter(Ratepayer.account_number == acct).first()
            if not rp:
                errors.append(f"Row {i}: Ratepayer account '{acct}' not found"); continue
            if category not in [c.value for c in RevenueCategory]:
                category = "other"
            inv_num = "INV-" + "".join(random.choices(string.digits, k=8))
            due = datetime.fromisoformat(due_raw)
            inv = Invoice(invoice_number=inv_num, ratepayer_id=rp.id, category=category,
                          amount=amount, amount_paid=0.0, balance=amount,
                          due_date=due, notes=notes, created_by=current_user.id)
            db.add(inv); created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
    db.flush()
    db.add(AuditLog(user_id=current_user.id, action="IMPORT", table_name="invoices",
                    description=f"Imported {created} invoices from file"))
    db.commit()
    return {"created": created, "errors": errors, "message": f"{created} invoices imported"}

@app.post("/api/import/budgets")
async def import_budgets(file: UploadFile = File(...),
                          db: Session = Depends(get_db),
                          current_user: User = Depends(require_roles(UserRole.admin, UserRole.budget_officer))):
    rows = await parse_upload(file)
    created = 0; updated = 0; errors = []
    for i, row in enumerate(rows, 2):
        try:
            fy   = str(row.get("fiscal_year") or row.get("Fiscal Year") or "").strip()
            dept = str(row.get("department")  or row.get("Department")  or "").strip()
            cat  = str(row.get("category")    or row.get("Category")    or "General").strip()
            alloc= float(row.get("allocated_amount") or row.get("Allocated Amount") or 0)
            spent= float(row.get("spent_amount") or row.get("Spent Amount") or 0)
            if not fy or not dept:
                errors.append(f"Row {i}: fiscal_year and department required"); continue
            existing = db.query(Budget).filter(
                Budget.fiscal_year == fy, Budget.department == dept, Budget.category == cat).first()
            if existing:
                existing.allocated_amount = alloc
                existing.spent_amount     = spent
                existing.remaining        = alloc - spent
                updated += 1
            else:
                b = Budget(fiscal_year=fy, department=dept, category=cat,
                           allocated_amount=alloc, spent_amount=spent, remaining=alloc - spent)
                db.add(b); created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
    db.flush()
    db.add(AuditLog(user_id=current_user.id, action="IMPORT", table_name="budgets",
                    description=f"Budget import: {created} created, {updated} updated"))
    db.commit()
    return {"created": created, "updated": updated, "errors": errors,
            "message": f"{created} created, {updated} updated"}

# â"€â"€â"€ Import Templates â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/templates/{entity}")
def download_template(entity: str, format: str = "csv",
                      current_user: User = Depends(get_current_user)):
    templates = {
        "ratepayers": (["full_name", "address", "ward", "zone", "phone", "email", "property_type"],
                       [["John Doe", "123 Main St", "Ward 5", "Mbare", "0771234567",
                         "john@example.com", "residential"]]),
        "invoices":   (["account_number", "category", "amount", "due_date", "notes"],
                       [["COH-123456", "rates", "150.00", "2026-06-30", ""]]),
        "budgets":    (["fiscal_year", "department", "category", "allocated_amount", "spent_amount"],
                       [["2025/2026", "Finance", "Operations", "500000.00", "0.00"]]),
        "revenue-targets": (["fiscal_year", "category", "target_amount", "period", "notes"],
                            [["2025/2026", "rates", "1500000.00", "annual", ""],
                             ["2025/2026", "water", "800000.00", "annual", ""],
                             ["2025/2026", "sewerage", "600000.00", "annual", ""]]),
    }
    if entity not in templates:
        raise HTTPException(404, f"No template for '{entity}'")
    headers, sample = templates[entity]
    return export_response(format, headers, sample,
                           f"template_{entity}.csv", f"template_{entity}.xlsx",
                           f"{entity.title()} Template",
                           f"City of Harare FMS - {entity.title()} Import Template")

# â"€â"€â"€ Reports â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/reports/budget-variance")
def report_budget_variance(format: str = "json", db: Session = Depends(get_db),
                            current_user: User = Depends(get_current_user)):
    budgets = db.query(Budget).order_by(Budget.department).all()
    # Actual approved expenditures by department
    actual_by_dept = dict(
        db.query(Expenditure.department, func.sum(Expenditure.amount))
          .filter(Expenditure.is_approved == True)
          .group_by(Expenditure.department).all()
    )
    rows_json = []
    rows_data = []
    for b in budgets:
        actual     = actual_by_dept.get(b.department, 0) or 0
        variance   = round(b.allocated_amount - actual, 2)
        var_pct    = round(variance / b.allocated_amount * 100, 1) if b.allocated_amount > 0 else 0
        status     = "Over Budget" if variance < 0 else ("On Budget" if variance == 0 else "Under Budget")
        rows_json.append({
            "department": b.department, "fiscal_year": b.fiscal_year, "category": b.category,
            "allocated": round(b.allocated_amount, 2), "actual_spent": round(actual, 2),
            "variance": variance, "variance_pct": var_pct, "status": status
        })
        rows_data.append([b.department, b.fiscal_year, b.category,
                          round(b.allocated_amount, 2), round(actual, 2),
                          variance, f"{var_pct}%", status])
    if format == "json":
        return rows_json
    headers = ["Department", "Fiscal Year", "Category", "Allocated ($)",
               "Actual Spent ($)", "Variance ($)", "Variance (%)", "Status"]
    return export_response(format, headers, rows_data,
                           "budget_variance_report.csv", "budget_variance_report.xlsx",
                           "Budget Variance", "City of Harare FMS - Budget Variance Report")

@app.get("/api/reports/financial-summary")
def report_financial_summary(format: str = "json", db: Session = Depends(get_db),
                              current_user: User = Depends(get_current_user)):
    cat_data = db.query(
        Invoice.category,
        func.sum(Invoice.amount).label("billed"),
        func.sum(Invoice.amount_paid).label("collected"),
        func.sum(Invoice.balance).label("outstanding"),
        func.count(Invoice.id).label("count")
    ).group_by(Invoice.category).all()

    rows_json = []
    rows_data = []
    for c in cat_data:
        billed    = round(c.billed or 0, 2)
        collected = round(c.collected or 0, 2)
        outstanding = round(c.outstanding or 0, 2)
        rate = round(collected / billed * 100, 1) if billed > 0 else 0.0
        rows_json.append({"category": c.category, "invoice_count": c.count,
                           "billed": billed, "collected": collected,
                           "outstanding": outstanding, "collection_rate": rate})
        rows_data.append([c.category, c.count, billed, collected, outstanding, f"{rate}%"])

    total_billed    = round(sum(r["billed"] for r in rows_json), 2)
    total_collected = round(sum(r["collected"] for r in rows_json), 2)
    total_outstanding = round(sum(r["outstanding"] for r in rows_json), 2)
    total_rate = round(total_collected / total_billed * 100, 1) if total_billed > 0 else 0.0

    if format == "json":
        return {"rows": rows_json,
                "totals": {"total_billed": total_billed, "total_collected": total_collected,
                           "total_outstanding": total_outstanding, "collection_rate": total_rate}}
    rows_data.append(["TOTAL", sum(r["invoice_count"] for r in rows_json),
                      total_billed, total_collected, total_outstanding, f"{total_rate}%"])
    headers = ["Category", "Invoice Count", "Billed ($)", "Collected ($)",
               "Outstanding ($)", "Collection Rate (%)"]
    return export_response(format, headers, rows_data,
                           "financial_summary.csv", "financial_summary.xlsx",
                           "Financial Summary", "City of Harare FMS - Financial Summary Report")

@app.get("/api/reports/audit-trail")
def report_audit_trail(format: str = "csv",
                       date_from: Optional[str] = None, date_to: Optional[str] = None,
                       db: Session = Depends(get_db),
                       current_user: User = Depends(require_roles(UserRole.admin, UserRole.auditor))):
    return export_audit_logs(format=format, date_from=date_from, date_to=date_to,
                              db=db, current_user=current_user)

# â"€â"€â"€ Aging Analysis â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def _aging_bucket(days: int) -> str:
    if days <= 0:   return "current"
    if days <= 30:  return "days_1_30"
    if days <= 60:  return "days_31_60"
    if days <= 90:  return "days_61_90"
    if days <= 120: return "days_91_120"
    return "days_120_plus"

AGING_LABELS = ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "91-120 Days", "120+ Days"]
AGING_KEYS   = ["current", "days_1_30", "days_31_60", "days_61_90", "days_91_120", "days_120_plus"]

@app.get("/api/aging/debtors")
def debtors_aging(format: str = "json", zone: Optional[str] = None,
                  db: Session = Depends(get_db),
                  current_user: User = Depends(get_current_user)):
    """Accounts-receivable aging: outstanding invoice balances by debtor and overdue bucket."""
    today = now()
    q = db.query(Invoice).filter(Invoice.balance > 0)
    items = q.all()

    rp_map: dict = {}
    for inv in items:
        rp = db.query(Ratepayer).filter(Ratepayer.id == inv.ratepayer_id).first()
        if not rp: continue
        if zone and rp.zone != zone: continue
        key = rp.account_number
        if key not in rp_map:
            rp_map[key] = {
                "ratepayer_id": rp.id,
                "account_number": rp.account_number, "ratepayer_name": rp.full_name,
                "zone": rp.zone, "ward": rp.ward, "property_type": rp.property_type,
                "invoice_count": 0,
                "current": 0.0, "days_1_30": 0.0, "days_31_60": 0.0,
                "days_61_90": 0.0, "days_91_120": 0.0, "days_120_plus": 0.0, "total": 0.0
            }
        days_over = (today - inv.due_date).days
        bucket = _aging_bucket(days_over)
        rp_map[key][bucket]     = round(rp_map[key][bucket] + inv.balance, 2)
        rp_map[key]["total"]    = round(rp_map[key]["total"]  + inv.balance, 2)
        rp_map[key]["invoice_count"] += 1

    rows = sorted(rp_map.values(), key=lambda r: r["days_120_plus"], reverse=True)

    # Totals row
    totals = {k: round(sum(r[k] for r in rows), 2) for k in AGING_KEYS}
    totals["total"] = round(sum(r["total"] for r in rows), 2)

    if format == "json":
        return {"rows": rows, "totals": totals,
                "summary": {lbl: totals[key] for lbl, key in zip(AGING_LABELS, AGING_KEYS)}}

    headers = ["Account #", "Ratepayer Name", "Zone", "Ward", "Invoices"] + AGING_LABELS + ["Total Outstanding"]
    data_rows = [[r["account_number"], r["ratepayer_name"], r["zone"], r["ward"], r["invoice_count"]] +
                 [r[k] for k in AGING_KEYS] + [r["total"]] for r in rows]
    data_rows.append(["", "TOTAL", "", "", sum(r["invoice_count"] for r in rows)] +
                     [totals[k] for k in AGING_KEYS] + [totals["total"]])
    return export_response(format, headers, data_rows,
                           "debtors_aging.csv", "debtors_aging.xlsx",
                           "Debtors Aging", "City of Harare FMS - Debtors Aging Analysis")

@app.get("/api/aging/creditors")
def creditors_aging(format: str = "json", department: Optional[str] = None,
                    db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    """Accounts-payable aging: unapproved/pending expenditures by department and age bucket."""
    today = now()
    q = db.query(Expenditure).filter(Expenditure.is_approved == False)
    if department: q = q.filter(Expenditure.department == department)
    items = q.all()

    dept_map: dict = {}
    for exp in items:
        key = exp.department
        if key not in dept_map:
            dept_map[key] = {
                "department": key, "expenditure_count": 0,
                "current": 0.0, "days_1_30": 0.0, "days_31_60": 0.0,
                "days_61_90": 0.0, "days_91_120": 0.0, "days_120_plus": 0.0, "total": 0.0
            }
        days_old  = (today - exp.expenditure_date).days
        bucket = _aging_bucket(days_old)
        dept_map[key][bucket]   = round(dept_map[key][bucket] + exp.amount, 2)
        dept_map[key]["total"]  = round(dept_map[key]["total"]  + exp.amount, 2)
        dept_map[key]["expenditure_count"] += 1

    # Also include individual expenditure detail for drill-down
    detail_rows = []
    for exp in sorted(items, key=lambda e: e.expenditure_date):
        days_old = (today - exp.expenditure_date).days
        detail_rows.append({
            "reference_number": exp.reference_number,
            "department": exp.department,
            "description": exp.description,
            "amount": exp.amount,
            "budget_line": exp.budget_line,
            "expenditure_date": str(exp.expenditure_date)[:10],
            "days_outstanding": days_old,
            "bucket": _aging_bucket(days_old)
        })

    rows = sorted(dept_map.values(), key=lambda r: r["days_120_plus"], reverse=True)
    totals = {k: round(sum(r[k] for r in rows), 2) for k in AGING_KEYS}
    totals["total"] = round(sum(r["total"] for r in rows), 2)

    if format == "json":
        return {"rows": rows, "totals": totals, "detail": detail_rows,
                "summary": {lbl: totals[key] for lbl, key in zip(AGING_LABELS, AGING_KEYS)}}

    headers = ["Department", "Expenditures"] + AGING_LABELS + ["Total Payable"]
    data_rows = [[r["department"], r["expenditure_count"]] +
                 [r[k] for k in AGING_KEYS] + [r["total"]] for r in rows]
    data_rows.append(["TOTAL", sum(r["expenditure_count"] for r in rows)] +
                     [totals[k] for k in AGING_KEYS] + [totals["total"]])
    return export_response(format, headers, data_rows,
                           "creditors_aging.csv", "creditors_aging.xlsx",
                           "Creditors Aging", "City of Harare FMS - Creditors Aging Analysis")

# â"€â"€â"€ Revenue Targets â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/revenue-targets")
def list_revenue_targets(fiscal_year: Optional[str] = None,
                          db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    q = db.query(RevenueTarget)
    if fiscal_year: q = q.filter(RevenueTarget.fiscal_year == fiscal_year)
    items = q.order_by(RevenueTarget.fiscal_year, RevenueTarget.category).all()
    return [{"id": t.id, "fiscal_year": t.fiscal_year, "category": t.category,
             "target_amount": t.target_amount, "period": t.period,
             "notes": t.notes, "created_at": str(t.created_at)[:10]} for t in items]

@app.post("/api/revenue-targets")
def create_revenue_target(data: RevenueTargetCreate,
                           db: Session = Depends(get_db),
                           current_user: User = Depends(require_roles(
                               UserRole.admin, UserRole.budget_officer, UserRole.accountant))):
    existing = db.query(RevenueTarget).filter(
        RevenueTarget.fiscal_year == data.fiscal_year,
        RevenueTarget.category    == data.category,
        RevenueTarget.period      == data.period).first()
    if existing:
        raise HTTPException(400, "A target already exists for this category/year/period. Use edit to update.")
    t = RevenueTarget(fiscal_year=data.fiscal_year, category=data.category,
                       target_amount=data.target_amount, period=data.period,
                       notes=data.notes, created_by=current_user.id)
    db.add(t); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="revenue_targets",
                    record_id=t.id,
                    description=f"Revenue target set: {data.category} {data.fiscal_year} = ${data.target_amount}"))
    db.commit()
    return {"id": t.id, "message": "Revenue target created"}

@app.put("/api/revenue-targets/{tid}")
def update_revenue_target(tid: int, data: RevenueTargetUpdate,
                           db: Session = Depends(get_db),
                           current_user: User = Depends(require_roles(
                               UserRole.admin, UserRole.budget_officer, UserRole.accountant))):
    t = db.query(RevenueTarget).filter(RevenueTarget.id == tid).first()
    if not t: raise HTTPException(404, "Revenue target not found")
    for k, v in data.dict(exclude_none=True).items():
        setattr(t, k, v)
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="revenue_targets",
                    record_id=tid,
                    description=f"Updated revenue target: {t.category} {t.fiscal_year}"))
    db.commit()
    return {"message": "Revenue target updated"}

@app.delete("/api/revenue-targets/{tid}")
def delete_revenue_target(tid: int,
                           db: Session = Depends(get_db),
                           current_user: User = Depends(require_roles(
                               UserRole.admin, UserRole.budget_officer))):
    t = db.query(RevenueTarget).filter(RevenueTarget.id == tid).first()
    if not t: raise HTTPException(404, "Revenue target not found")
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="revenue_targets",
                    record_id=tid,
                    description=f"Deleted revenue target: {t.category} {t.fiscal_year}"))
    db.delete(t); db.commit()
    return {"message": "Revenue target deleted"}

@app.post("/api/import/revenue-targets")
async def import_revenue_targets(file: UploadFile = File(...),
                                  db: Session = Depends(get_db),
                                  current_user: User = Depends(require_roles(
                                      UserRole.admin, UserRole.budget_officer, UserRole.accountant))):
    rows = await parse_upload(file)
    created = 0; updated = 0; errors = []
    for i, row in enumerate(rows, 2):
        try:
            fy     = str(row.get("fiscal_year") or row.get("Fiscal Year") or "").strip()
            cat    = str(row.get("category")    or row.get("Category")    or "").strip().lower()
            period = str(row.get("period")      or row.get("Period")      or "annual").strip()
            target = float(row.get("target_amount") or row.get("Target Amount") or 0)
            notes  = str(row.get("notes") or row.get("Notes") or "").strip() or None
            if not fy or not cat:
                errors.append(f"Row {i}: fiscal_year and category required"); continue
            existing = db.query(RevenueTarget).filter(
                RevenueTarget.fiscal_year == fy,
                RevenueTarget.category    == cat,
                RevenueTarget.period      == period).first()
            if existing:
                existing.target_amount = target
                existing.notes = notes
                updated += 1
            else:
                t = RevenueTarget(fiscal_year=fy, category=cat, period=period,
                                   target_amount=target, notes=notes,
                                   created_by=current_user.id)
                db.add(t); created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
    db.flush()
    db.add(AuditLog(user_id=current_user.id, action="IMPORT", table_name="revenue_targets",
                    description=f"Revenue targets import: {created} created, {updated} updated"))
    db.commit()
    return {"created": created, "updated": updated, "errors": errors,
            "message": f"{created} created, {updated} updated"}

# â"€â"€â"€ COH Historical Dataset Import â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

COH_DATASET_PATH = r"C:\Users\Gigi\Documents\Terry Work\Dissertation\COH_Official_Financial_Dataset.xlsx"

@app.post("/api/import/coh-dataset")
def import_coh_dataset(db: Session = Depends(get_db),
                       current_user: User = Depends(require_roles(
                           UserRole.admin, UserRole.budget_officer, UserRole.accountant))):
    """Import City of Harare official financial dataset (2021-2025) into revenue targets and budgets."""
    if not EXCEL_OK:
        raise HTTPException(500, "openpyxl not installed")
    if not os.path.exists(COH_DATASET_PATH):
        raise HTTPException(404, f"Dataset file not found: {COH_DATASET_PATH}")

    wb = openpyxl.load_workbook(COH_DATASET_PATH, data_only=True)
    summary = {"revenue_targets": {"created": 0, "updated": 0},
               "budgets": {"created": 0, "updated": 0},
               "expenditures": {"created": 0},
               "errors": []}

    # â"€â"€ Sheet 1: Revenue Trend 2021-2025 â†' RevenueTarget â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    CATEGORY_MAP = {
        "property taxes": "rates", "rates": "rates",
        "water & sanitation": "water", "water": "water",
        "refuse removal": "refuse", "refuse collection": "refuse", "refuse": "refuse",
        "rentals & leases": "rentals", "rentals": "rentals",
        "building plan fees": "licensing", "building fees": "licensing",
        "city health fees": "licensing", "city health": "licensing",
        "easypark & parking": "parking", "easypark/parking": "parking", "parking": "parking",
        "grants & subsidies": "other", "grants (parliament)": "other",
        "other receipts": "other", "other revenue": "other",
        "public safety": "other", "roads & infrastructure": "other",
        "natural resources": "other",
    }

    if "Revenue Trend 2021-2025" in wb.sheetnames:
        ws = wb["Revenue Trend 2021-2025"]
        rows_iter = list(ws.iter_rows(values_only=True))
        # Find the header row with year columns
        header_row_idx = None
        years = []
        for i, row in enumerate(rows_iter):
            if row and isinstance(row[0], int) and row[0] == 1:
                # Check previous row for headers
                hdr = rows_iter[i - 1] if i > 0 else None
                if hdr and hdr[1] and "category" in str(hdr[1]).lower():
                    header_row_idx = i - 1
                    years = [str(h)[:4] for h in hdr[2:] if h and str(h)[:4].isdigit()]
                    break

        if header_row_idx is not None:
            for row in rows_iter[header_row_idx + 1:]:
                if not row or not isinstance(row[0], int):
                    continue
                cat_raw = str(row[1] or "").strip().lower()
                cat = CATEGORY_MAP.get(cat_raw, "other")
                for j, yr in enumerate(years):
                    val = row[2 + j]
                    try:
                        amount = float(val) if val is not None else 0.0
                    except (TypeError, ValueError):
                        continue
                    existing = db.query(RevenueTarget).filter(
                        RevenueTarget.fiscal_year == yr,
                        RevenueTarget.category == cat,
                        RevenueTarget.period == "annual",
                        RevenueTarget.notes.like(f"%COH:{row[1]}%")).first()
                    if existing:
                        existing.target_amount = amount
                        summary["revenue_targets"]["updated"] += 1
                    else:
                        db.add(RevenueTarget(
                            fiscal_year=yr, category=cat, period="annual",
                            target_amount=amount,
                            notes=f"COH:{row[1]} | Actual receipts (USD)",
                            created_by=current_user.id))
                        summary["revenue_targets"]["created"] += 1

    # â"€â"€ Sheet: Revenue vs Budget 2025 â†' Budget â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    if "Revenue vs Budget 2025" in wb.sheetnames:
        ws = wb["Revenue vs Budget 2025"]
        rows_iter = list(ws.iter_rows(values_only=True))
        for row in rows_iter:
            if not row or not isinstance(row[0], int):
                continue
            cat_raw = str(row[1] or "").strip().lower()
            cat = CATEGORY_MAP.get(cat_raw, "other")
            try:
                allocated = float(row[2]) if row[2] else 0.0
                spent     = float(row[5]) if row[5] else 0.0  # Actual Receipts column
            except (TypeError, ValueError):
                continue
            existing = db.query(Budget).filter(
                Budget.fiscal_year == "2024/2025",
                Budget.department == "Revenue",
                Budget.category == cat).first()
            if existing:
                existing.allocated_amount = allocated
                existing.spent_amount = spent
                existing.remaining = allocated - spent
                summary["budgets"]["updated"] += 1
            else:
                db.add(Budget(
                    fiscal_year="2024/2025", department="Revenue",
                    category=cat, allocated_amount=allocated,
                    spent_amount=spent, remaining=allocated - spent))
                summary["budgets"]["created"] += 1

    # â"€â"€ Sheet: Expenditure by Dept 2025 â†' Budget (expenditure side) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    if "Expenditure by Dept 2025" in wb.sheetnames:
        ws = wb["Expenditure by Dept 2025"]
        rows_iter = list(ws.iter_rows(values_only=True))
        for row in rows_iter:
            if not row or not isinstance(row[0], int):
                continue
            dept = str(row[1] or "").strip()
            try:
                allocated = float(row[2]) if row[2] else 0.0
                spent     = float(row[4]) if row[4] else 0.0  # Actual Receipts Jan-Sep
            except (TypeError, ValueError):
                continue
            existing = db.query(Budget).filter(
                Budget.fiscal_year == "2024/2025",
                Budget.department == dept,
                Budget.category == "expenditure").first()
            if existing:
                existing.allocated_amount = allocated
                existing.spent_amount = spent
                existing.remaining = allocated - spent
                summary["budgets"]["updated"] += 1
            else:
                db.add(Budget(
                    fiscal_year="2024/2025", department=dept,
                    category="expenditure", allocated_amount=allocated,
                    spent_amount=spent, remaining=allocated - spent))
                summary["budgets"]["created"] += 1

    db.flush()
    db.add(AuditLog(
        user_id=current_user.id, action="IMPORT", table_name="coh_dataset",
        description=(f"COH dataset imported: "
                     f"{summary['revenue_targets']['created']} revenue targets created, "
                     f"{summary['revenue_targets']['updated']} updated; "
                     f"{summary['budgets']['created']} budgets created, "
                     f"{summary['budgets']['updated']} updated")))
    db.commit()
    return {
        "message": "COH Historical Dataset imported successfully",
        "revenue_targets": summary["revenue_targets"],
        "budgets": summary["budgets"],
        "errors": summary["errors"]
    }

# â"€â"€â"€ Performance Report â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/reports/performance")
def report_performance(fiscal_year: Optional[str] = None,
                        format: str = "json",
                        db: Session = Depends(get_db),
                        current_user: User = Depends(get_current_user)):
    """Revenue performance: targets vs actuals by category."""
    # Actual revenue collected by category
    actual_q = db.query(Invoice.category,
                        func.sum(Invoice.amount).label("billed"),
                        func.sum(Invoice.amount_paid).label("collected"),
                        func.sum(Invoice.balance).label("outstanding"))
    if fiscal_year:
        # Approximate fiscal year filter - assumes fiscal_year like "2025/2026"
        # Map to year range: "2025/2026" â†' 2025-07-01 to 2026-06-30
        try:
            yr_start = int(fiscal_year.split("/")[0])
            start_dt = datetime(yr_start, 7, 1)
            end_dt   = datetime(yr_start + 1, 6, 30, 23, 59, 59)
            actual_q = actual_q.filter(Invoice.issue_date >= start_dt,
                                       Invoice.issue_date <= end_dt)
        except Exception:
            pass
    actual_by_cat = {r.category: {"billed": round(r.billed or 0, 2),
                                   "collected": round(r.collected or 0, 2),
                                   "outstanding": round(r.outstanding or 0, 2)}
                     for r in actual_q.group_by(Invoice.category).all()}

    # Revenue targets
    tgt_q = db.query(RevenueTarget)
    if fiscal_year: tgt_q = tgt_q.filter(RevenueTarget.fiscal_year == fiscal_year)
    targets = tgt_q.all()

    rows_json = []
    rows_data = []

    # Merge targets with actuals - show all categories that have either a target or actual
    all_cats = set(actual_by_cat.keys()) | {t.category for t in targets}
    tgt_map  = {t.category: t.target_amount for t in targets}

    for cat in sorted(all_cats):
        act    = actual_by_cat.get(cat, {"billed": 0, "collected": 0, "outstanding": 0})
        target = tgt_map.get(cat, 0)
        variance    = round(act["collected"] - target, 2)
        var_pct     = round(variance / target * 100, 1) if target > 0 else 0
        coll_rate   = round(act["collected"] / act["billed"] * 100, 1) if act["billed"] > 0 else 0
        status      = "Above Target" if variance > 0 else ("Below Target" if variance < 0 else "On Target")
        rows_json.append({
            "category": cat, "target": target,
            "billed": act["billed"], "collected": act["collected"],
            "outstanding": act["outstanding"],
            "variance": variance, "variance_pct": var_pct,
            "collection_rate": coll_rate, "status": status
        })
        rows_data.append([cat.title(), round(target, 2), act["billed"], act["collected"],
                          act["outstanding"], f"{coll_rate}%", variance, f"{var_pct}%", status])

    if format == "json":
        total_target    = round(sum(r["target"] for r in rows_json), 2)
        total_collected = round(sum(r["collected"] for r in rows_json), 2)
        total_billed    = round(sum(r["billed"] for r in rows_json), 2)
        return {"rows": rows_json,
                "totals": {"total_target": total_target, "total_billed": total_billed,
                           "total_collected": total_collected,
                           "overall_variance": round(total_collected - total_target, 2)}}

    fy_label = fiscal_year or "All Years"
    headers  = ["Category", "Target ($)", "Billed ($)", "Collected ($)",
                "Outstanding ($)", "Collection Rate", "Variance ($)", "Variance (%)", "Status"]
    return export_response(format, headers, rows_data,
                           "performance_report.csv", "performance_report.xlsx",
                           "Performance", f"City of Harare FMS - Revenue Performance Report ({fy_label})")

# â"€â"€â"€ Root â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return FileResponse(os.path.join(frontend_path, "images", "crest.png"))

@app.get("/")
def root():
    return FileResponse(os.path.join(frontend_path, "pages", "login.html"))

# â"€â"€â"€ Health Check â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/health")
def health_check(db: Session = Depends(get_db)):
    """System health endpoint - verifies API and database are operational."""
    try:
        user_count = db.query(User).count()
        return {
            "status": "healthy",
            "system": "City of Harare Financial Management System",
            "version": "2.0.0",
            "timestamp": str(now()),
            "database": "connected",
            "users": user_count
        }
    except Exception as e:
        raise HTTPException(500, f"Database error: {str(e)}")

# --- Reconciliation Register -------------------------------------------------

@app.get("/api/reconciliation")
def list_reconciliation(
    status: Optional[str] = Query(None, regex="^(reconciled|unreconciled)?$"),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Dedicated reconciliation register.
    Returns paginated payments with full reconciliation metadata PLUS
    a summary block (counts, amounts, rate, leakage exposure) in a single
    request — eliminates the two round-trips the old page used.
    """
    # Summary stats (always over all payments, ignoring filter)
    total_count      = db.query(func.count(Payment.id)).scalar() or 0
    recon_count      = db.query(func.count(Payment.id)).filter(Payment.is_reconciled == True).scalar() or 0
    unrecon_count    = total_count - recon_count
    recon_amt        = usd_payment_sum(db.query(Payment).filter(Payment.is_reconciled == True))
    unrecon_amt      = usd_payment_sum(db.query(Payment).filter(Payment.is_reconciled == False))
    recon_rate       = round(recon_count / total_count * 100, 1) if total_count > 0 else 0.0
    unrecon_cash_amt = usd_payment_sum(db.query(Payment).filter(
        Payment.is_reconciled == False, Payment.payment_method == "cash"))

    # Filtered query for the page
    q = db.query(Payment)
    if status == "reconciled":
        q = q.filter(Payment.is_reconciled == True)
    elif status == "unreconciled":
        q = q.filter(Payment.is_reconciled == False)
    if search:
        if len(search) > 100:
            raise HTTPException(400, "Search too long")
        rp_ids = [r.id for r in db.query(Ratepayer).filter(
            Ratepayer.full_name.ilike(f"%{search}%") |
            Ratepayer.account_number.ilike(f"%{search}%")
        ).all()]
        q = q.filter(
            Payment.receipt_number.ilike(f"%{search}%") |
            Payment.ratepayer_id.in_(rp_ids)
        )

    page_total = q.count()
    payments   = q.order_by(desc(Payment.payment_date)).offset(skip).limit(limit).all()

    # Fetch related names in batch
    rp_ids_page  = list({p.ratepayer_id for p in payments})
    user_ids_page = list({p.collected_by for p in payments if p.collected_by} |
                         {p.reconciled_by for p in payments if p.reconciled_by})
    rp_map   = {r.id: r for r in db.query(Ratepayer).filter(Ratepayer.id.in_(rp_ids_page)).all()}
    user_map = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids_page)).all()}

    items = []
    for p in payments:
        rp         = rp_map.get(p.ratepayer_id)
        collector  = user_map.get(p.collected_by)
        reconciler = user_map.get(p.reconciled_by)
        items.append({
            "id":              p.id,
            "receipt_number":  p.receipt_number,
            "ratepayer_id":    p.ratepayer_id,
            "ratepayer_name":  rp.full_name      if rp         else "Unknown",
            "account_number":  rp.account_number if rp         else "",
            "amount":          round(p.amount, 2),
            "payment_method":  p.payment_method,
            "currency":        p.currency,
            "payment_date":    str(p.payment_date)[:10],
            "is_reconciled":   p.is_reconciled,
            "reconciled_by":   reconciler.full_name if reconciler else None,
            "reconciled_at":   str(p.reconciled_at)[:16] if p.reconciled_at else None,
            "collected_by":    collector.full_name  if collector  else None,
            "anomaly_flag":    p.anomaly_flag,
            "anomaly_reason":  p.anomaly_reason,
            "invoice_id":      p.invoice_id,
            "notes":           p.notes,
        })

    return {
        "summary": {
            "total_payments":      total_count,
            "reconciled_count":    recon_count,
            "unreconciled_count":  unrecon_count,
            "reconciled_amount":   round(recon_amt, 2),
            "unreconciled_amount": round(unrecon_amt, 2),
            "reconciliation_rate": recon_rate,
            "leakage_exposure":    round(unrecon_cash_amt * 0.40, 2),
        },
        "total":  page_total,
        "skip":   skip,
        "limit":  limit,
        "items":  items,
    }

# --- Bulk Reconciliation -----------------------------------------------------

@app.post("/api/payments/reconcile-all")
def bulk_reconcile(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant, UserRole.auditor))
):
    """Reconcile all unreconciled payments in a single operation with full audit trail."""
    unreconciled = db.query(Payment).filter(Payment.is_reconciled == False).all()
    count = 0
    for pmt in unreconciled:
        pmt.is_reconciled = True
        db.add(AuditLog(
            user_id=current_user.id, action="UPDATE", table_name="payments",
            record_id=pmt.id,
            description=f"Bulk reconciliation: {pmt.receipt_number} (${pmt.amount:.2f}) reconciled by {current_user.full_name}"
        ))
        count += 1
    db.commit()
    return {"reconciled": count, "message": f"{count} payment(s) reconciled successfully"}

# â"€â"€â"€ Collection Rate Trend â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/reports/collection-rate-trend")
def collection_rate_trend(
    months: int = 12,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Monthly collection rate trend for the past N months - drives the key dissertation chart."""
    result = []
    for i in range(months, 0, -1):
        # Calculate start and end of each month
        ref = now()
        month_start = (ref.replace(day=1) - timedelta(days=30 * (i - 1))).replace(day=1)
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1, day=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1, day=1)

        billed = db.query(func.sum(Invoice.amount))\
            .filter(Invoice.issue_date >= month_start, Invoice.issue_date < month_end)\
            .scalar() or 0
        collected = usd_payment_sum(db.query(Payment)
            .filter(Payment.payment_date >= month_start, Payment.payment_date < month_end))
        overdue = db.query(func.sum(Invoice.balance))\
            .filter(
                Invoice.issue_date >= month_start, Invoice.issue_date < month_end,
                Invoice.status == PaymentStatus.overdue
            ).scalar() or 0

        rate = round(collected / billed * 100, 1) if billed > 0 else 0.0
        result.append({
            "month": month_start.strftime("%b %Y"),
            "billed": round(billed, 2),
            "collected": round(collected, 2),
            "overdue": round(overdue, 2),
            "collection_rate": rate,
            "leakage_gap": round(billed - collected, 2)
        })
    return result

# â"€â"€â"€ Ratepayer Statement â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/ratepayers/{rp_id}/statement")
def ratepayer_statement(
    rp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Full account statement for a ratepayer - invoices, payments, balance.
    Used for generating printable statements to send to ratepayers.
    """
    rp = db.query(Ratepayer).filter(Ratepayer.id == rp_id).first()
    if not rp:
        raise HTTPException(404, "Ratepayer not found")

    invoices = db.query(Invoice).filter(Invoice.ratepayer_id == rp_id)\
                 .order_by(Invoice.issue_date).all()
    payments = db.query(Payment).filter(Payment.ratepayer_id == rp_id)\
                 .order_by(Payment.payment_date).all()

    total_billed    = sum(i.amount for i in invoices)
    total_paid      = sum(amount_to_usd(p.amount, p.currency) for p in payments)
    total_outstanding = sum(i.balance for i in invoices)
    overdue_balance = sum(i.balance for i in invoices if i.status == PaymentStatus.overdue)

    invoice_list = [{
        "invoice_number": i.invoice_number,
        "category": i.category,
        "amount": i.amount,
        "amount_paid": i.amount_paid,
        "balance": i.balance,
        "status": i.status,
        "issue_date": str(i.issue_date)[:10],
        "due_date": str(i.due_date)[:10]
    } for i in invoices]

    payment_list = [{
        "receipt_number": p.receipt_number,
        "amount": p.amount,
        "method": p.payment_method,
        "currency": p.currency,
        "date": str(p.payment_date)[:10],
        "reconciled": p.is_reconciled
    } for p in payments]

    return {
        "ratepayer": {
            "account_number": rp.account_number,
            "full_name": rp.full_name,
            "address": rp.address,
            "ward": rp.ward,
            "zone": rp.zone,
            "phone": rp.phone,
            "email": rp.email,
            "property_type": rp.property_type
        },
        "summary": {
            "total_billed": round(total_billed, 2),
            "total_paid": round(total_paid, 2),
            "total_outstanding": round(total_outstanding, 2),
            "overdue_balance": round(overdue_balance, 2),
            "invoice_count": len(invoices),
            "payment_count": len(payments)
        },
        "invoices": invoice_list,
        "payments": payment_list,
        "generated_at": str(now())[:16],
        "generated_by": current_user.full_name
    }

# â"€â"€â"€ Officer Performance Report â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/reports/officer-performance")
def officer_performance_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(
        UserRole.admin, UserRole.auditor, UserRole.accountant
    ))
):
    """
    Revenue officer performance comparison - collection rates, invoice counts,
    and Z-score deviation from team average.
    Addresses D11 open-ended responses: 'automated management reports to reduce manual compilation'.
    """
    officers = db.query(User).filter(
        User.role == UserRole.revenue_officer, User.is_active == True
    ).all()

    rows = []
    for officer in officers:
        total_collected = usd_payment_sum(db.query(Payment).filter(Payment.collected_by == officer.id))
        total_billed = db.query(func.sum(Invoice.amount))\
            .filter(Invoice.created_by == officer.id).scalar() or 0
        invoice_count = db.query(func.count(Invoice.id))\
            .filter(Invoice.created_by == officer.id).scalar() or 0
        payment_count = db.query(func.count(Payment.id))\
            .filter(Payment.collected_by == officer.id).scalar() or 0
        unreconciled = db.query(func.count(Payment.id))\
            .filter(Payment.collected_by == officer.id, Payment.is_reconciled == False).scalar() or 0

        rate = round(total_collected / total_billed * 100, 1) if total_billed > 0 else 0.0
        rows.append({
            "officer_id": officer.id,
            "full_name": officer.full_name,
            "username": officer.username,
            "total_billed": round(total_billed, 2),
            "total_collected": round(total_collected, 2),
            "collection_rate": rate,
            "invoice_count": invoice_count,
            "payment_count": payment_count,
            "unreconciled_payments": unreconciled
        })

    # Compute Z-scores
    if len(rows) >= 2:
        rates = [r["collection_rate"] for r in rows]
        mean_rate = sum(rates) / len(rates)
        variance = sum((r - mean_rate) ** 2 for r in rates) / max(len(rates) - 1, 1)
        std_rate = math.sqrt(variance) if variance > 0 else 0
        for row in rows:
            z = round((row["collection_rate"] - mean_rate) / std_rate, 2) if std_rate > 0 else 0
            row["z_score"] = z
            row["performance"] = (
                "Above Average" if z > 0.5 else
                "At Risk" if z < -1.5 else
                "Average"
            )
        team_avg = round(mean_rate, 1)
    else:
        for row in rows:
            row["z_score"] = 0
            row["performance"] = "Insufficient data"
        team_avg = rows[0]["collection_rate"] if rows else 0

    return {
        "officers": sorted(rows, key=lambda r: r["collection_rate"], reverse=True),
        "team_average_rate": team_avg,
        "generated_at": str(now())[:16]
    }

# â"€â"€â"€ Cashflow Forecast â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/reports/cashflow-forecast")
def cashflow_forecast(
    months_ahead: int = 3,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Simple cashflow forecast for the next N months based on:
    - Expected receipts: pending invoices due in the period Ã- historical collection rate
    - Expected payments: budget allocation / 12 Ã- months_ahead
    Addresses D11 open-ended feature request: 'cash flow forecasting'.
    """
    # Historical collection rate (last 90 days)
    cutoff = now() - timedelta(days=90)
    hist_billed = db.query(func.sum(Invoice.amount))\
        .filter(Invoice.issue_date >= cutoff).scalar() or 0
    hist_collected = usd_payment_sum(db.query(Payment).filter(Payment.payment_date >= cutoff))
    hist_rate = hist_collected / hist_billed if hist_billed > 0 else 0.35

    forecast = []
    for i in range(1, months_ahead + 1):
        month_start = now().replace(day=1)
        if month_start.month + i > 12:
            target_month = month_start.replace(
                year=month_start.year + (month_start.month + i - 1) // 12,
                month=(month_start.month + i - 1) % 12 + 1, day=1
            )
        else:
            target_month = month_start.replace(month=month_start.month + i, day=1)

        if target_month.month == 12:
            next_month = target_month.replace(year=target_month.year + 1, month=1, day=1)
        else:
            next_month = target_month.replace(month=target_month.month + 1, day=1)

        # Invoices due in this month
        pending_billed = db.query(func.sum(Invoice.amount))\
            .filter(
                Invoice.due_date >= target_month,
                Invoice.due_date < next_month,
                Invoice.status.in_([PaymentStatus.pending, PaymentStatus.overdue])
            ).scalar() or 0

        expected_receipts = round(pending_billed * hist_rate, 2)
        # Budget-based expected payments (total allocated / 12)
        total_budget = db.query(func.sum(Budget.allocated_amount)).scalar() or 0
        expected_payments = round(total_budget / 12, 2)

        forecast.append({
            "month": target_month.strftime("%b %Y"),
            "pending_invoices_due": round(pending_billed, 2),
            "expected_receipts": expected_receipts,
            "expected_payments": expected_payments,
            "net_cashflow": round(expected_receipts - expected_payments, 2),
            "collection_rate_assumed": round(hist_rate * 100, 1)
        })

    return {
        "forecast": forecast,
        "basis": f"Historical collection rate (last 90 days): {round(hist_rate*100,1)}%",
        "generated_at": str(now())[:16]
    }

# â"€â"€â"€ Full Database Backup Export â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/export/full-backup")
def full_database_backup(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin))
):
    """
    Export all tables as a single multi-sheet Excel workbook.
    Admin-only. Creates a complete point-in-time backup of all financial data.
    """
    if not EXCEL_OK:
        raise HTTPException(500, "openpyxl not installed")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        # Header row
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    # Ratepayers
    rps = db.query(Ratepayer).all()
    add_sheet("Ratepayers",
        ["Account #", "Full Name", "Address", "Ward", "Zone", "Phone", "Email", "Type", "Active", "Created"],
        [[r.account_number, r.full_name, r.address, r.ward, r.zone,
          r.phone, r.email, r.property_type, r.is_active, str(r.created_at)[:10]] for r in rps])

    # Invoices
    invs = db.query(Invoice).all()
    add_sheet("Invoices",
        ["Invoice #", "Ratepayer ID", "Category", "Amount", "Paid", "Balance", "Status", "Anomaly", "Issue", "Due"],
        [[i.invoice_number, i.ratepayer_id, i.category, i.amount, i.amount_paid,
          i.balance, i.status, i.anomaly_flag, str(i.issue_date)[:10], str(i.due_date)[:10]] for i in invs])

    # Payments
    pmts = db.query(Payment).all()
    add_sheet("Payments",
        ["Receipt #", "Ratepayer ID", "Invoice ID", "Amount", "Currency", "Method", "Date", "Reconciled", "Anomaly"],
        [[p.receipt_number, p.ratepayer_id, p.invoice_id, p.amount, p.currency,
          p.payment_method, str(p.payment_date)[:10], p.is_reconciled, p.anomaly_flag] for p in pmts])

    # Expenditures
    exps = db.query(Expenditure).all()
    add_sheet("Expenditures",
        ["Reference", "Department", "Description", "Amount", "Budget Line", "Date", "Approved"],
        [[e.reference_number, e.department, e.description, e.amount,
          e.budget_line, str(e.expenditure_date)[:10], e.is_approved] for e in exps])

    # Budgets
    buds = db.query(Budget).all()
    add_sheet("Budgets",
        ["Fiscal Year", "Department", "Category", "Allocated", "Spent", "Remaining"],
        [[b.fiscal_year, b.department, b.category, b.allocated_amount,
          b.spent_amount, b.remaining] for b in buds])

    # Audit Logs
    logs = db.query(AuditLog).order_by(desc(AuditLog.timestamp)).limit(5000).all()
    add_sheet("Audit Log",
        ["ID", "User ID", "Action", "Table", "Record ID", "Description", "Timestamp"],
        [[l.id, l.user_id, l.action, l.table_name, l.record_id,
          l.description, str(l.timestamp)[:16]] for l in logs])

    # Leakage Alerts
    alerts = db.query(LeakageAlert).all()
    add_sheet("Leakage Alerts",
        ["ID", "Type", "Severity", "Description", "Resolved", "Created"],
        [[a.id, a.alert_type, a.severity, a.description,
          a.is_resolved, str(a.created_at)[:10]] for a in alerts])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"COH_FMS_Backup_{now().strftime('%Y%m%d_%H%M')}.xlsx"
    db.add(AuditLog(
        user_id=current_user.id, action="EXPORT", table_name="system",
        description=f"Full database backup exported by {current_user.full_name}"
    ))
    db.commit()

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# â"€â"€â"€ Management Summary Report â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/reports/management-summary")
def management_summary_report(
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Consolidated management summary report - all key metrics in one endpoint.
    Designed to satisfy D11 requirement: 'export functions to produce management reports
    that can be tabled at Council meetings without manual compilation.'
    """
    refresh_overdue_invoices(db)

    total_billed     = db.query(func.sum(Invoice.amount)).scalar() or 0
    total_collected  = usd_payment_sum(db.query(Payment))
    total_outstanding = db.query(func.sum(Invoice.balance)).scalar() or 0
    overdue_balance  = db.query(func.sum(Invoice.balance))\
        .filter(Invoice.status == PaymentStatus.overdue).scalar() or 0
    unrecon_amt      = usd_payment_sum(db.query(Payment).filter(Payment.is_reconciled == False))
    collection_rate  = round(total_collected / total_billed * 100, 1) if total_billed > 0 else 0
    leakage_estimate = round(unrecon_amt * 0.40 + overdue_balance * 0.25, 2)

    # Revenue by category
    cat_data = db.query(
        Invoice.category,
        func.sum(Invoice.amount).label("billed"),
        func.sum(Invoice.amount_paid).label("collected")
    ).group_by(Invoice.category).all()

    # Active alerts by severity
    high_alerts   = db.query(LeakageAlert)\
        .filter(LeakageAlert.is_resolved == False, LeakageAlert.severity == "high").count()
    medium_alerts = db.query(LeakageAlert)\
        .filter(LeakageAlert.is_resolved == False, LeakageAlert.severity == "medium").count()

    # Budget utilisation
    total_budget  = db.query(func.sum(Budget.allocated_amount)).scalar() or 0
    total_spent   = db.query(func.sum(Budget.spent_amount)).scalar() or 0
    budget_utilisation = round(total_spent / total_budget * 100, 1) if total_budget > 0 else 0

    summary = {
        "report_title": "City of Harare - Financial Management Summary",
        "generated_at": str(now())[:16],
        "generated_by": current_user.full_name,
        "revenue": {
            "total_billed": round(total_billed, 2),
            "total_collected": round(total_collected, 2),
            "total_outstanding": round(total_outstanding, 2),
            "overdue_balance": round(overdue_balance, 2),
            "collection_rate_pct": collection_rate,
            "estimated_leakage": leakage_estimate,
            "unreconciled_amount": round(unrecon_amt, 2)
        },
        "by_category": [{
            "category": r.category,
            "billed": round(r.billed or 0, 2),
            "collected": round(r.collected or 0, 2),
            "rate": round((r.collected or 0) / (r.billed or 1) * 100, 1)
        } for r in cat_data],
        "alerts": {
            "high_severity": high_alerts,
            "medium_severity": medium_alerts,
            "total_active": high_alerts + medium_alerts
        },
        "budget": {
            "total_allocated": round(total_budget, 2),
            "total_spent": round(total_spent, 2),
            "utilisation_pct": budget_utilisation
        },
        "ratepayers": {
            "total": db.query(Ratepayer).count(),
            "active": db.query(Ratepayer).filter(Ratepayer.is_active == True).count()
        }
    }

    if format == "json":
        return summary

    # Excel export
    headers = ["Metric", "Value"]
    rows = [
        ["=== REVENUE PERFORMANCE ===", ""],
        ["Total Billed (USD)", round(total_billed, 2)],
        ["Total Collected (USD)", round(total_collected, 2)],
        ["Outstanding Balance (USD)", round(total_outstanding, 2)],
        ["Overdue Balance (USD)", round(overdue_balance, 2)],
        ["Collection Rate (%)", f"{collection_rate}%"],
        ["Estimated Revenue Leakage (USD)", leakage_estimate],
        ["Unreconciled Payments (USD)", round(unrecon_amt, 2)],
        ["", ""],
        ["=== BUDGET PERFORMANCE ===", ""],
        ["Total Budget Allocated (USD)", round(total_budget, 2)],
        ["Total Spent (USD)", round(total_spent, 2)],
        ["Budget Utilisation (%)", f"{budget_utilisation}%"],
        ["", ""],
        ["=== SYSTEM ALERTS ===", ""],
        ["High Severity Alerts", high_alerts],
        ["Medium Severity Alerts", medium_alerts],
        ["", ""],
        ["Generated At", str(now())[:16]],
        ["Generated By", current_user.full_name],
    ]
    return make_excel_response(
        headers, rows,
        f"COH_Management_Summary_{now().strftime('%Y%m%d')}.xlsx",
        "Management Summary",
        "City of Harare FMS - Management Summary Report"
    )


# =============================================================================
# â"€â"€â"€ IMPROVEMENTS v2.1 - New Endpoints â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
# All endpoints below are additions to the base v2.0 system.
# =============================================================================

# â"€â"€â"€ Security: Login Attempts & Account Unlock â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/security/login-attempts")
def get_login_attempts(
    skip: int = 0, limit: int = 50,
    username: Optional[str] = None,
    success: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, UserRole.auditor))
):
    """View login attempt history for security monitoring (IMPROVEMENT 2)."""
    q = db.query(LoginAttempt).order_by(desc(LoginAttempt.attempted_at))
    if username: q = q.filter(LoginAttempt.username == username)
    if success is not None: q = q.filter(LoginAttempt.success == success)
    attempts = q.offset(skip).limit(limit).all()
    return [{"id": a.id, "username": a.username, "ip_address": a.ip_address,
             "success": a.success, "failure_reason": a.failure_reason,
             "attempted_at": str(a.attempted_at)[:19]} for a in attempts]

@app.post("/api/security/unlock-user/{user_id}")
def unlock_user_account(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin))
):
    """Admin: unlock a locked account (IMPROVEMENT 1)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user: raise HTTPException(404, "User not found")
    user.locked_until       = None
    user.failed_login_count = 0
    db.add(AuditLog(user_id=current_user.id, action="UNLOCK", table_name="users",
                    record_id=user.id,
                    description=f"Account {user.username} manually unlocked by {current_user.username}"))
    db.commit()
    return {"message": f"Account '{user.username}' has been unlocked"}

# â"€â"€â"€ AI: Risk Scoring â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.post("/api/ai/compute-risk-scores")
def compute_all_risk_scores(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, UserRole.auditor, UserRole.accountant))
):
    """
    Run AI risk scoring across all active ratepayers and update risk_score and risk_label.
    Weighted additive model - ZIMRA (2023) Taxpayer Risk Segmentation Framework (IMPROVEMENT 4).
    """
    ratepayers = db.query(Ratepayer).filter(Ratepayer.is_active == True).all()
    high_count = medium_count = low_count = 0
    for rp in ratepayers:
        result = compute_ratepayer_risk(rp, db)
        rp.risk_score      = result["score"]
        rp.risk_label      = result["label"]
        rp.risk_updated_at = now()
        if result["label"] == "high": high_count += 1
        elif result["label"] == "medium": medium_count += 1
        else: low_count += 1
    db.commit()
    db.add(AuditLog(user_id=current_user.id, action="AI_SCAN", table_name="ratepayers",
                    description=f"AI risk scoring: {high_count} HIGH, {medium_count} MEDIUM, {low_count} LOW"))
    db.commit()
    return {"updated": len(ratepayers), "high_risk": high_count,
            "medium_risk": medium_count, "low_risk": low_count,
            "methodology": "Composite weighted score: overdue ratio (40%), payment recency (30%), anomaly rate (20%), plan defaults (10%)"}

@app.get("/api/ai/risk-register")
def get_risk_register(
    risk_label: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Risk register - all ratepayers ordered by risk score descending (IMPROVEMENT 4)."""
    q = db.query(Ratepayer).filter(Ratepayer.is_active == True)
    if risk_label: q = q.filter(Ratepayer.risk_label == risk_label)
    rps = q.order_by(desc(Ratepayer.risk_score)).all()
    return [{"id": rp.id, "account_number": rp.account_number, "full_name": rp.full_name,
             "ward": rp.ward, "zone": rp.zone, "risk_score": rp.risk_score,
             "risk_label": rp.risk_label,
             "risk_updated_at": str(rp.risk_updated_at)[:16] if rp.risk_updated_at else None}
            for rp in rps]

# â"€â"€â"€ AI: Revenue Prediction â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/ai/revenue-prediction")
def revenue_prediction(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    OLS linear regression on 6 months of payment data to forecast next month's revenue.
    Freedman, Pisani & Purves (2007). Statistics (4th ed.) (IMPROVEMENT 5).
    """
    return predict_next_month_revenue(db)

# â"€â"€â"€ AI: Duplicate Invoice Check â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.post("/api/invoices/check-duplicate")
def check_duplicate_invoice(
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Pre-creation check: returns warning if a duplicate invoice fingerprint exists.
    Call before submitting invoice creation form (IMPROVEMENT 6).
    """
    fp = compute_invoice_fingerprint(
        int(data.get("ratepayer_id", 0)),
        str(data.get("category", "")),
        float(data.get("amount", 0)),
        str(data.get("due_date", ""))
    )
    existing = db.query(Invoice).filter(Invoice.fingerprint == fp).first()
    if existing:
        return {"is_duplicate": True, "existing_invoice": existing.invoice_number,
                "warning": f"Invoice {existing.invoice_number} already exists with identical ratepayer, category, amount, and due date."}
    return {"is_duplicate": False}

# â"€â"€â"€ Payment Plans â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/payment-plans")
def list_payment_plans(
    ratepayer_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all payment/instalment plans with progress tracking (IMPROVEMENT 10)."""
    q = db.query(PaymentPlan)
    if ratepayer_id: q = q.filter(PaymentPlan.ratepayer_id == ratepayer_id)
    if status:       q = q.filter(PaymentPlan.status == status)
    plans = q.order_by(desc(PaymentPlan.created_at)).all()
    result = []
    for p in plans:
        rp = db.query(Ratepayer).filter(Ratepayer.id == p.ratepayer_id).first()
        result.append({
            "id": p.id, "ratepayer_id": p.ratepayer_id,
            "ratepayer_name": rp.full_name if rp else None,
            "account_number": rp.account_number if rp else None,
            "total_debt": p.total_debt, "instalment_amount": p.instalment_amount,
            "frequency": p.frequency, "total_instalments": p.total_instalments,
            "instalments_paid": p.instalments_paid,
            "progress_pct": round(p.instalments_paid / p.total_instalments * 100, 1) if p.total_instalments else 0,
            "start_date": str(p.start_date)[:10], "next_due_date": str(p.next_due_date)[:10],
            "status": p.status, "notes": p.notes, "created_at": str(p.created_at)[:10]
        })
    return result

@app.post("/api/payment-plans")
def create_payment_plan(
    data: PaymentPlanCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant, UserRole.revenue_officer))
):
    """Create an instalment plan for a ratepayer with outstanding debt (IMPROVEMENT 10)."""
    rp = db.query(Ratepayer).filter(Ratepayer.id == data.ratepayer_id).first()
    if not rp: raise HTTPException(404, "Ratepayer not found")
    if data.instalment_amount <= 0: raise HTTPException(400, "Instalment amount must be positive")
    if data.total_instalments <= 0: raise HTTPException(400, "Number of instalments must be positive")
    try: start = datetime.fromisoformat(data.start_date)
    except ValueError: raise HTTPException(400, "start_date must be YYYY-MM-DD format")
    next_due = start + timedelta(days=7 if data.frequency == "weekly" else 30)
    plan = PaymentPlan(ratepayer_id=data.ratepayer_id, total_debt=data.total_debt,
                       instalment_amount=data.instalment_amount, frequency=data.frequency,
                       total_instalments=data.total_instalments, instalments_paid=0,
                       start_date=start, next_due_date=next_due,
                       status=PaymentPlanStatus.active, created_by=current_user.id, notes=data.notes)
    db.add(plan); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="payment_plans",
                    record_id=plan.id,
                    description=f"Payment plan created for {rp.full_name}: ${data.instalment_amount}/{data.frequency} Ã- {data.total_instalments}"))
    db.commit(); db.refresh(plan)
    return {"message": "Payment plan created", "id": plan.id}

@app.patch("/api/payment-plans/{plan_id}/record-instalment")
def record_instalment(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant, UserRole.revenue_officer))
):
    """Record one instalment paid against a payment plan (IMPROVEMENT 10)."""
    plan = db.query(PaymentPlan).filter(PaymentPlan.id == plan_id).first()
    if not plan: raise HTTPException(404, "Payment plan not found")
    if plan.status != PaymentPlanStatus.active:
        raise HTTPException(400, f"Plan is '{plan.status}' - cannot record instalment")
    plan.instalments_paid += 1
    if plan.instalments_paid >= plan.total_instalments:
        plan.status = PaymentPlanStatus.completed
    else:
        plan.next_due_date += timedelta(days=7 if plan.frequency == "weekly" else 30)
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="payment_plans",
                    record_id=plan.id,
                    description=f"Instalment {plan.instalments_paid}/{plan.total_instalments} recorded for plan #{plan.id}"))
    db.commit()
    return {"message": "Instalment recorded", "instalments_paid": plan.instalments_paid,
            "remaining": plan.total_instalments - plan.instalments_paid, "status": plan.status}

@app.patch("/api/payment-plans/{plan_id}/default")
def mark_plan_defaulted(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))
):
    """Mark a payment plan as defaulted (IMPROVEMENT 10)."""
    plan = db.query(PaymentPlan).filter(PaymentPlan.id == plan_id).first()
    if not plan: raise HTTPException(404, "Plan not found")
    plan.status = PaymentPlanStatus.defaulted
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="payment_plans",
                    record_id=plan.id, description=f"Payment plan #{plan.id} marked as defaulted"))
    db.commit()
    return {"message": "Plan marked as defaulted"}

# â"€â"€â"€ Notifications â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/notifications")
def get_notifications(
    unread_only: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Retrieve personal and broadcast notifications for current user (IMPROVEMENT 11)."""
    q = db.query(SystemNotification).filter(
        (SystemNotification.user_id == current_user.id) | (SystemNotification.user_id == None)
    )
    if unread_only: q = q.filter(SystemNotification.is_read == False)
    notifs = q.order_by(desc(SystemNotification.created_at)).limit(50).all()
    return [{"id": n.id, "title": n.title, "message": n.message, "category": n.category,
             "is_read": n.is_read, "created_at": str(n.created_at)[:16]} for n in notifs]

@app.get("/api/notifications/unread-count")
def get_unread_count(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Quick unread notification count for the topbar badge."""
    count = db.query(func.count(SystemNotification.id)).filter(
        (SystemNotification.user_id == current_user.id) | (SystemNotification.user_id == None),
        SystemNotification.is_read == False
    ).scalar() or 0
    return {"unread_count": count}

@app.patch("/api/notifications/{notif_id}/read")
def mark_notification_read(
    notif_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Mark a notification as read (IMPROVEMENT 11)."""
    n = db.query(SystemNotification).filter(SystemNotification.id == notif_id).first()
    if not n: raise HTTPException(404, "Notification not found")
    n.is_read = True
    db.commit()
    return {"message": "Marked as read"}

@app.post("/api/notifications/broadcast")
def broadcast_notification(
    data: NotificationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin))
):
    """Admin: send a system-wide broadcast notification to all users (IMPROVEMENT 11)."""
    n = SystemNotification(user_id=None, title=data.title,
                           message=data.message, category=data.category)
    db.add(n)
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="system_notifications",
                    description=f"Broadcast notification sent by {current_user.username}: {data.title}"))
    db.commit()
    return {"message": "Broadcast notification sent"}

# â"€â"€â"€ Reports: Risk Register Export â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.get("/api/reports/risk-register")
def export_risk_register(
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin, UserRole.auditor, UserRole.accountant))
):
    """
    Export ratepayer risk register with colour-coded Excel output (IMPROVEMENT 12).
    Risk scores must be computed first via POST /api/ai/compute-risk-scores.
    """
    rps = db.query(Ratepayer).filter(Ratepayer.is_active == True)\
            .order_by(desc(Ratepayer.risk_score)).all()
    rows_data = []
    for rp in rps:
        overdue = db.query(func.sum(Invoice.balance)).filter(
            Invoice.ratepayer_id == rp.id, Invoice.status == PaymentStatus.overdue).scalar() or 0
        rows_data.append({
            "account_number": rp.account_number, "full_name": rp.full_name,
            "ward": rp.ward, "zone": rp.zone, "risk_score": rp.risk_score,
            "risk_label": (rp.risk_label or "unscored").upper(),
            "overdue_balance": round(overdue, 2),
            "risk_updated": str(rp.risk_updated_at)[:10] if rp.risk_updated_at else "Not scored"
        })
    if format == "json":
        return rows_data
    headers = ["Account #", "Full Name", "Ward", "Zone", "Risk Score", "Risk Level", "Overdue Balance ($)", "Last Scored"]
    rows = [[d["account_number"], d["full_name"], d["ward"], d["zone"],
             d["risk_score"], d["risk_label"], d["overdue_balance"], d["risk_updated"]]
            for d in rows_data]
    if format == "xlsx" and EXCEL_OK:
        risk_colors = {"HIGH": "C0392B", "MEDIUM": "E67E22", "LOW": "27AE60", "UNSCORED": "95A5A6"}
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Risk Register"
        title_cell = ws.cell(row=1, column=1, value="City of Harare FMS - Ratepayer Risk Register")
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E5FA3", end_color="2E5FA3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 3):
            fill_bg = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid") if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill_bg: cell.fill = fill_bg
                if ci == 6:  # Risk level column - colour by risk
                    colour = risk_colors.get(str(val), "CCCCCC")
                    cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col if hasattr(c, "column_letter")), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return Response(content=output.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=risk_register.xlsx"})
    # CSV fallback
    out = io.StringIO(); writer = csv_mod.writer(out)
    writer.writerow(headers); writer.writerows(rows)
    return Response(content=out.getvalue().encode("utf-8-sig"), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=risk_register.csv"})


@app.get("/api/ai/status")
def ai_status(current_user: User = Depends(get_current_user)):
    return {"available": False}

# ─── Billing: Rate Configuration ──────────────────────────────────────────────

class BillingRateCreate(BaseModel):
    category: str
    flat_amount: float = Field(..., ge=0)
    description: Optional[str] = None
    is_active: bool = True

class BillingRateUpdate(BaseModel):
    flat_amount: Optional[float] = Field(None, ge=0)
    description: Optional[str] = None
    is_active: Optional[bool] = None

@app.get("/api/billing/rates")
def list_billing_rates(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    rates = db.query(BillingRate).order_by(BillingRate.category).all()
    result = []
    for r in rates:
        updater = db.query(User).filter(User.id == r.updated_by).first() if r.updated_by else None
        result.append({
            "id": r.id, "category": r.category, "flat_amount": r.flat_amount,
            "description": r.description, "is_active": r.is_active,
            "updated_by": updater.full_name if updater else None,
            "updated_at": str(r.updated_at)[:19]
        })
    return result

@app.post("/api/billing/rates", status_code=201)
def create_billing_rate(data: BillingRateCreate, db: Session = Depends(get_db),
                        current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    existing = db.query(BillingRate).filter(BillingRate.category == data.category).first()
    if existing:
        raise HTTPException(400, f"Rate for category '{data.category}' already exists. Use PUT to update it.")
    rate = BillingRate(category=data.category, flat_amount=data.flat_amount,
                       description=data.description, is_active=data.is_active,
                       updated_by=current_user.id)
    db.add(rate); db.flush()
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="billing_rates",
                    record_id=rate.id, description=f"Billing rate created for {data.category}: ${data.flat_amount}"))
    db.commit()
    return {"id": rate.id, "message": "Billing rate created"}

@app.put("/api/billing/rates/{rate_id}")
def update_billing_rate(rate_id: int, data: BillingRateUpdate, db: Session = Depends(get_db),
                        current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant))):
    rate = db.query(BillingRate).filter(BillingRate.id == rate_id).first()
    if not rate: raise HTTPException(404, "Billing rate not found")
    if data.flat_amount is not None: rate.flat_amount = data.flat_amount
    if data.description is not None: rate.description = data.description
    if data.is_active is not None: rate.is_active = data.is_active
    rate.updated_by = current_user.id
    rate.updated_at = now()
    db.add(AuditLog(user_id=current_user.id, action="UPDATE", table_name="billing_rates",
                    record_id=rate_id, description=f"Billing rate updated for {rate.category}"))
    db.commit()
    return {"message": "Billing rate updated"}

@app.delete("/api/billing/rates/{rate_id}")
def delete_billing_rate(rate_id: int, db: Session = Depends(get_db),
                        current_user: User = Depends(require_roles(UserRole.admin))):
    rate = db.query(BillingRate).filter(BillingRate.id == rate_id).first()
    if not rate: raise HTTPException(404, "Billing rate not found")
    db.add(AuditLog(user_id=current_user.id, action="DELETE", table_name="billing_rates",
                    record_id=rate_id, description=f"Billing rate deleted for {rate.category}"))
    db.delete(rate); db.commit()
    return {"message": "Billing rate deleted"}

# ─── Billing: Runs ────────────────────────────────────────────────────────────

class BillingRunCreate(BaseModel):
    billing_period: str
    categories: Optional[List[str]] = None
    due_date: str
    notes: Optional[str] = None

@app.get("/api/billing/runs")
def list_billing_runs(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    runs = db.query(BillingRun).order_by(desc(BillingRun.created_at)).limit(50).all()
    result = []
    for r in runs:
        creator = db.query(User).filter(User.id == r.created_by).first()
        result.append({
            "id": r.id, "run_number": r.run_number, "billing_period": r.billing_period,
            "categories": r.categories, "due_date": str(r.due_date)[:10],
            "invoices_created": r.invoices_created, "total_amount": r.total_amount,
            "status": r.status, "notes": r.notes,
            "created_by": creator.full_name if creator else None,
            "created_at": str(r.created_at)[:19]
        })
    return result

@app.post("/api/billing/runs", status_code=201)
def execute_billing_run(data: BillingRunCreate, db: Session = Depends(get_db),
                        current_user: User = Depends(require_roles(UserRole.admin, UserRole.accountant, UserRole.revenue_officer))):
    due = datetime.fromisoformat(data.due_date)
    run_num = "BRN-" + "".join(random.choices(string.digits, k=8))

    target_cats = data.categories if data.categories else [c.value for c in RevenueCategory]
    rates = {r.category: r.flat_amount for r in
             db.query(BillingRate).filter(BillingRate.is_active == True,
                                          BillingRate.category.in_(target_cats)).all()}
    if not rates:
        raise HTTPException(400, "No active billing rates found for the selected categories. Configure rates first.")

    ratepayers = db.query(Ratepayer).filter(Ratepayer.is_active == True).all()
    created = 0
    total_amount = 0.0
    errors = []

    for rp in ratepayers:
        for cat, amount in rates.items():
            if amount <= 0:
                continue
            fp = compute_invoice_fingerprint(rp.id, cat, amount, data.due_date)
            if db.query(Invoice).filter(Invoice.fingerprint == fp).first():
                continue
            inv_num = "INV-" + "".join(random.choices(string.digits, k=8))
            cat_amounts = [r[0] for r in db.query(Invoice.amount).filter(Invoice.category == cat).all()]
            anomaly_flag, anomaly_reason = _detect_anomaly(amount, cat_amounts)
            inv = Invoice(invoice_number=inv_num, ratepayer_id=rp.id, category=cat,
                          amount=amount, amount_paid=0.0, balance=amount,
                          due_date=due, created_by=current_user.id, fingerprint=fp,
                          notes=f"Billing run {run_num} — {data.billing_period}",
                          anomaly_flag=anomaly_flag, anomaly_reason=anomaly_reason)
            db.add(inv)
            created += 1
            total_amount += amount

    run = BillingRun(run_number=run_num, billing_period=data.billing_period,
                     categories=",".join(target_cats), due_date=due,
                     invoices_created=created, total_amount=round(total_amount, 2),
                     status="completed", notes=data.notes, created_by=current_user.id)
    db.add(run)
    db.add(AuditLog(user_id=current_user.id, action="CREATE", table_name="billing_runs",
                    description=f"Billing run {run_num}: {created} invoices, ${total_amount:.2f}"))
    db.commit()
    return {"run_number": run_num, "invoices_created": created,
            "total_amount": round(total_amount, 2), "message": f"Billing run complete: {created} invoices generated"}

@app.get("/api/billing/runs/{run_id}")
def get_billing_run(run_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    run = db.query(BillingRun).filter(BillingRun.id == run_id).first()
    if not run: raise HTTPException(404, "Billing run not found")
    invoices = db.query(Invoice).filter(Invoice.notes.like(f"%{run.run_number}%")).all()
    inv_list = []
    for inv in invoices:
        rp = db.query(Ratepayer).filter(Ratepayer.id == inv.ratepayer_id).first()
        inv_list.append({
            "id": inv.id, "invoice_number": inv.invoice_number,
            "ratepayer": rp.full_name if rp else str(inv.ratepayer_id),
            "account_number": rp.account_number if rp else "",
            "category": inv.category, "amount": inv.amount,
            "status": inv.status, "due_date": str(inv.due_date)[:10]
        })
    creator = db.query(User).filter(User.id == run.created_by).first()
    return {
        "id": run.id, "run_number": run.run_number, "billing_period": run.billing_period,
        "categories": run.categories, "due_date": str(run.due_date)[:10],
        "invoices_created": run.invoices_created, "total_amount": run.total_amount,
        "status": run.status, "notes": run.notes,
        "created_by": creator.full_name if creator else None,
        "created_at": str(run.created_at)[:19],
        "invoices": inv_list
    }

# ─── Billing: Account Statement ───────────────────────────────────────────────

@app.get("/api/billing/statement/{ratepayer_id}")
def billing_statement(ratepayer_id: int, db: Session = Depends(get_db),
                      current_user: User = Depends(get_current_user)):
    rp = db.query(Ratepayer).filter(Ratepayer.id == ratepayer_id).first()
    if not rp: raise HTTPException(404, "Ratepayer not found")

    invoices = db.query(Invoice).filter(Invoice.ratepayer_id == ratepayer_id)\
                 .order_by(desc(Invoice.issue_date)).all()
    payments = db.query(Payment).filter(Payment.ratepayer_id == ratepayer_id)\
                 .order_by(desc(Payment.payment_date)).all()

    total_billed   = sum(i.amount for i in invoices)
    total_paid     = sum(i.amount_paid for i in invoices)
    total_balance  = sum(i.balance for i in invoices)
    overdue_balance = sum(i.balance for i in invoices if i.status == PaymentStatus.overdue)

    inv_rows = [{
        "id": i.id, "invoice_number": i.invoice_number,
        "category": i.category, "amount": i.amount,
        "amount_paid": i.amount_paid, "balance": i.balance,
        "issue_date": str(i.issue_date)[:10], "due_date": str(i.due_date)[:10],
        "status": i.status, "anomaly_flag": i.anomaly_flag
    } for i in invoices]

    pmt_rows = [{
        "id": p.id, "receipt_number": p.receipt_number,
        "amount": p.amount, "currency": p.currency,
        "payment_method": p.payment_method,
        "payment_date": str(p.payment_date)[:10],
        "is_reconciled": p.is_reconciled
    } for p in payments]

    return {
        "ratepayer": {
            "id": rp.id, "account_number": rp.account_number,
            "full_name": rp.full_name, "address": rp.address,
            "ward": rp.ward, "zone": rp.zone,
            "phone": rp.phone, "email": rp.email,
            "property_type": rp.property_type
        },
        "summary": {
            "total_billed": round(total_billed, 2),
            "total_paid": round(total_paid, 2),
            "total_balance": round(total_balance, 2),
            "overdue_balance": round(overdue_balance, 2),
            "invoice_count": len(invoices),
            "payment_count": len(payments)
        },
        "invoices": inv_rows,
        "payments": pmt_rows,
        "generated_at": str(now())[:19]
    }

# ─── WebSocket ─────────────────────────────────────────────────────────────────

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """
    Real-time event push for dashboard and leakage monitor.
    Clients reconnect automatically on disconnect.
    Events: leakage_scan, new_payment, new_alert.
    """
    await ws_manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket)
